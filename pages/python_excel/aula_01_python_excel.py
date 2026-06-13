# --- Importar as bibliotecas --- #
from PIL import Image
import streamlit as st

# --- Configurações iniciais da página --- #
st.set_page_config(
    page_title='Python para Excel - Aula 01',
    page_icon=Image.open('./assets/logo/logo.png'),
    layout='wide'
)

# --- Carregar o estilo das fontes --- #
with open('./assets/css/style.css', 'r') as css:
    st.html(f'<style>{css.read()}</style>')

# --- Colocar o título da aula --- #
st.html('<h1 class="fonte_titulo_aula">Aula 01: Criando seu Primeiro Arquivo Excel com openpyxl!</h1>')

# --- Vídeo --- #
with st.expander('Se quiser acompanhar com o vídeo, acesse aqui! 👇'):
        st.video('https://youtu.be/N1nsnorupxY')

# --- Código da aula --- #
st.subheader('Se quiser acessar o código completo da aula, clique [aqui](https://github.com/GTL98/canal_mundo_python/blob/main/Curso%20Completo%20de%20Python%20para%20Excel%3A%20Do%20Zero%20ao%20Especialista/Aula%2001/aula_01.ipynb)')
st.divider()

# --- Introdução --- #
st.subheader('E fala, devs! Tudo bem com vocês? Espero que sim!')
st.html('<h1 class="fonte_titulo_aula">Introdução</h1>')
st.html('''<p class="fonte_texto">A automação de planilhas eletrônicas representa um pilar fundamental na 
engenharia de dados e na otimização de fluxos de trabalho corporativos. No ecossistema de desenvolvimento 
da linguagem Python, a biblioteca <span class='texto_python'>openpyxl</span> destaca-se como a 
ferramenta padrão para a criação, edição e leitura de arquivos que seguem o formato aberto OpenXML, 
amplamente conhecidos pela extensão <span class='texto_python'>.xlsx</span>. Esta aula aborda de 
maneira aprofundada os conceitos necessários para inicializar ambientes, manipular pastas de trabalho 
diretamente na memória RAM, inserir dados sob diferentes paradigmas e persistir informações de forma 
segura no disco físico.</p>''')
st.subheader('Então sem mais delongas, bora para a aula!')
st.divider()

# --- Configuração do Ambiente e Instalação do Pacote --- #
st.html('<h1 class="fonte_titulo_aula">Configuração do Ambiente e Instalação do Pacote</h1>')
st.html('''<p class="fonte_texto">Antes de iniciar a estruturação lógica para a manipulação de dados, é 
necessário certificar-se de que as dependências externas estejam devidamente integradas ao ambiente de 
execução do Python. A distribuição padrão do Python não inclui o pacote de manipulação OpenXML 
nativamente, o que exige a sua instalação prévia. Você pode executar o seguinte comando no 
prompt de comando (CMD) ou na célula do Notebook Jupyter:</p>''')
st.code('pip install openpyxl', line_numbers=True)
st.html('''<p class="fonte_texto">Esse comando realiza o download do pacote a partir do repositório 
oficial (PyPI), verificando compatibilidades e instalando todas as dependências internas exigidas para 
que o interpretador Python consiga decodificar a estrutura XML compactada que compõe um arquivo Excel. 
A execução desse processo é o pré-requisito indispensável para qualquer rotina subsequente de automação 
de planilhas.</p>''')
st.divider()

# --- Inicialização do Workbook de Trabalho na Memória RAM --- #
st.html('<h1 class="fonte_titulo_aula">Inicialização do Workbook de Trabalho na Memória RAM</h1>')
st.html('''<p class="fonte_texto">Com o pacote instalado com sucesso, o desenvolvimento da aplicação 
prossegue com a importação da biblioteca e a criação de uma pasta de trabalho vazia, conceitualmente 
denominada de <i>Workbook</i>:</p>''')
st.code('''# --- Importar a biblioteca --- #
import openpyxl

# --- Criar o workbook na memória --- #
wb = openpyxl.Workbook()''', line_numbers=True)
st.html('''<p class="fonte_texto">O comando 
<span class='palavras_reservadas'>import</span> <span class='texto_python'>openpyxl</span> carrega o 
<i>namespace</i> da biblioteca para o escopo do script, permitindo o acesso às suas classes e funções 
nativas. Em seguida, a instrução <span class='texto_python'>wb = openpyxl.Workbook()</span> instancia 
um objeto da classe principal <span class='texto_python'>Workbook</span>.</p>''')
st.html('''<p class="fonte_texto">O comportamento desse construtor é de extrema relevância para a 
performance da aplicação: ele cria toda a representação estrutural da planilha diretamente na memória 
volátil (RAM) do computador. Isso significa que nenhuma alteração ou escrita parcial gera concorrência de 
leitura e escrita (I/O) no disco físico do computador até que um comando explícito de salvamento seja 
executado.</p>''')
st.html('''<p class="fonte_texto">Ademais, no momento de sua criação, o Workbook adota um modelo de 
alocação de memória sob demanda. Embora ele venha estruturado com uma planilha padrão, nenhuma célula 
física é criada na memória de forma imediata. O <span class='texto_python'>openpyxl</span> adota essa 
abordagem eficiente para evitar o desperdício de recursos computacionais; as células são geradas e 
alocadas na memória somente quando são acessadas ou modificadas pela primeira vez no fluxo 
de código.</p>''')
st.divider()

# --- Ativação e Identificação da Planilha de Trabalho --- #
st.html('<h1 class="fonte_titulo_aula">Ativação e Identificação da Planilha de Trabalho</h1>')
st.html('''<p class="fonte_texto">Um arquivo do Excel pode conter múltiplas guias ou abas para organizar 
diferentes conjuntos de dados. No momento da instanciação de um novo documento, o 
<span class='texto_python'>openpyxl</span> assegura a existência de pelo menos uma aba padrão ativa 
para que o arquivo não seja nulo:</p>''')
st.code("""# --- Obter/ativar a planilha padrão do arquivo --- #
ws = wb.active

# --- Título da planilha --- #
ws.title = 'Relatório de Vendas'""", line_numbers=True)
st.html('''<p class="fonte_texto">A linha de código <span class='texto_python'>ws = wb.active</span> 
utiliza a propriedade <span class='texto_python'>active</span> para capturar a planilha que está 
atualmente selecionada como ativa no contexto do <i>Workbook</i>, armazenando-a na variável 
<span class='texto_python'>ws</span> (uma convenção para <i>worksheet</i>). Sem essa captura de 
referência, as instruções de manipulação de dados não teriam um destino definido dentro da estrutura 
tridimensional do arquivo.</p>''')
st.html('''<p class="fonte_texto">A planilha criada automaticamente recebe, por padrão, o nome genérico de 
<b>Sheet</b> ou <b>Planilha 1</b>. Para fins de organização de dados corporativos e manutenção de 
padrões profissionais, o atributo <span class='texto_python'>ws.title</span> permite renomear essa aba. 
Ao atribuir o valor <span class='variaveis'>'Relatório de Vendas'</span> à propriedade 
<span class='texto_python'>ws.title</span>, garantimos que a interface do usuário final 
exiba de forma autoexplicativa a finalidade das tabelas ali contidas.</p>''')
st.divider()

# --- Métodos para Escrita de Dados --- #
st.html('<h1 class="fonte_titulo_aula">Métodos para Escrita de Dados</h1>')
st.html('''<p class="fonte_texto">A manipulação de valores no 
<span class='texto_python'>openpyxl</span> pode ser realizada seguindo duas filosofias distintas de 
endereçamento: a atribuição explícita por meio de notação de dicionário e a atribuição dinâmica 
por meio de coordenadas numéricas.</p>''')

# --- Escrita Explícita com Notação Alfanumérica --- #
st.html('<h2 class="fonte_subtitulo_aula">Escrita Explícita com Notação Alfanumérica</h2>')
st.html('''<p class="fonte_texto">A notação de endereçamento alfanumérico utiliza o padrão tradicional 
de contagem, onde as colunas são representadas por letras e as linhas por números 
inteiros positivos iniciados em 1:</p>''')
st.code("""# --- Escrita de informação em células de modo explícito --- #
ws['A1'] = 'ID Venda'
ws['B1'] = 'Valor Bruto'""", line_numbers=True)
st.html('''<p class="fonte_texto">Ao executar <span class='texto_python'>ws[</span>
<span class='variaveis'>'A1'</span><span class='texto_python'>] = </span>
<span class='variaveis'>'ID Venda'</span>, o interpretador Python utiliza o método interno de obtenção 
de itens para verificar se a célula <b>A1</b> já existe na estrutura de dados residente na memória RAM. 
Se ela não existir, o framework a instancia dinamicamente e armazena a cadeia de caracteres atribuída. 
Esse método de acesso explícito é excelente para a criação manual de cabeçalhos e para a definição de 
células isoladas com finalidades específicas, proporcionando um código limpo e de leitura imediata para 
quem analisa o script.</p>''')

# --- Escrita Dinâmica com Coordenadas Numéricas --- #
st.html('<h2 class="fonte_subtitulo_aula">Escrita Dinâmica com Coordenadas Numéricas</h2>')
st.html('''<p class="fonte_texto">Embora o endereçamento <b>A1</b> seja altamente legível, ele impõe 
barreiras no desenvolvimento de rotinas de automação complexas, nas quais os dados precisam ser 
inseridos de forma iterativa por meio de laços de repetição. Para resolver essa limitação, a classe da 
planilha disponibiliza o método <span class='texto_python'>.cell()</b>:</p>''')
st.code("""# --- Escrita de informação em células por coordenadas --- #
ws.cell(row=2, column=1, value=1001)
ws.cell(row=2, column=2, value=450.75)""", line_numbers=True)
st.html('''<p class="fonte_texto">O método <span class='texto_python'>.cell()</span> recebe parâmetros 
explicitamente nomeados para determinar a localização espacial exata da informação:</p>''')
st.html('<ul class="fonte_texto">'
        '<li>O argumento <span class="texto_python">row</span> aceita um número inteiro que indica a '
        'linha desejada, em uma base indexada em 1.</li>'
        '<li>O argumento <span class="texto_python">column</span> aceita um número inteiro correspondente '
        'à coluna, eliminando a necessidade de converter índices matemáticos em letras (onde a coluna 1 '
        'mapeia para A, a coluna 2 para B, e assim sucessivamente).</li>'
        '<li>O argumento <span class="texto_python">value</span> recebe o conteúdo a ser colocado na '
        'interseção dessas duas coordenadas, suportando tipos numéricos nativos como inteiros e números '
        'de ponto flutuante.</li>'
        '</ul>')
st.html('''<p class="fonte_texto">No código acima, a instrução 
<span class='texto_python'>ws.cell(row=</span><span class='numeros'>2</span>
<span class='texto_python'>, column=</span><span class='numeros'>1</span>
<span class='texto_python'>, value=</span><span class='numeros'>1001</span>
<span class='texto_python'>)</span> cria a célula na segunda linha da primeira coluna (<b>A2</b>) com o 
valor numérico inteiro <span class='numeros'>1001</span>. Já a instrução seguinte escreve o decimal 
<span class='numeros'>450.75</span> na segunda linha da segunda coluna (<b>B2</b>).</p>''')
st.html('''<p class="fonte_texto">É fundamental destacar que loops de varredura que apenas consultam 
coordenadas sem preencher valores devem ser estruturados com cautela. Devido ao modelo de instanciação 
dinâmica de objetos do <span class='texto_python'>openpyxl</span>, navegar cegamente por um intervalo 
de coordenadas criará milhares de objetos de células vazias na memória RAM, o que pode impactar 
negativamente a performance de servidores e rotinas de big data.</p>''')
st.html('''<p class="fonte_texto">Para estruturar e sintetizar as diferenças fundamentais entre essas 
abordagens de inserção de dados, veja o comparativo a seguir:</p>''')
st.html("""<style type="text/css">
.tg  {border-collapse:collapse;border-spacing:0;}
.tg td{border-color:black;border-style:solid;border-width:1px;font-family:Arial, sans-serif;font-size:14px;
  overflow:hidden;padding:10px 5px;word-break:normal;}
.tg th{border-color:black;border-style:solid;border-width:1px;font-family:Arial, sans-serif;font-size:14px;
  font-weight:normal;overflow:hidden;padding:10px 5px;word-break:normal;}
.tg .tg-4rfe{border-color:#000000;color:#1F1F1F;font-family:Arial, Helvetica, sans-serif !important;font-size:16px;font-weight:bold;
  text-align:center;vertical-align:top}
.tg .tg-lz2v{border-color:#000000;color:#1F1F1F;font-family:Arial, Helvetica, sans-serif !important;font-size:16px;
  text-align:center;vertical-align:top}
</style>
<table class="tg"><thead>
  <tr>
    <th class="tg-4rfe">Dimensão de Comparação</th>
    <th class="tg-4rfe">Escrita Explicita (dicionário)</th>
    <th class="tg-4rfe">Escrita por Coordenadas (<span class='texto_python'>.cell()</span>)</th>
  </tr></thead>
<tbody>
  <tr>
    <td class="tg-lz2v">Sintaxe de Acesso</td>
    <td class="tg-lz2v"><span class='texto_python'>ws[</span><span class='variaveis'>'A1'</span>
    <span class='texto_python'>]</span></td>
    <td class="tg-lz2v"><span class='texto_python'>ws.cell(row=</span>
    <span class='numeros'>1</span><span class='texto_python'>, column=</span>
    <span class='numeros'>1</span><span class='texto_python'>)</span></td>
  </tr>
  <tr>
    <td class="tg-lz2v">Identificador de Coluna</td>
    <td class="tg-lz2v">Caractere alfabético</td>
    <td class="tg-lz2v">Inteiro de base</td>
  </tr>
  <tr>
    <td class="tg-lz2v">Indexação Temporal</td>
    <td class="tg-lz2v">Estática e manual</td>
    <td class="tg-lz2v">Dinâmica e propramátiga</td>
  </tr>
  <tr>
    <td class="tg-lz2v">Caso de Uso Ideal</td>
    <td class="tg-lz2v">Cabeçalhos e células de metadados únicos</td>
    <td class="tg-lz2v">Alimentação automatizada via loops e matrizes</td>
  </tr>
  <tr>
    <td class="tg-lz2v">Tipos de Dados Aceitos</td>
    <td class="tg-lz2v">Strings, Inteiros e Decimais</td>
    <td class="tg-lz2v">Strings, Inteiros e Decimais</td>
  </tr>
</tbody></table>""")
st.divider()

# --- Persistência de Dados e Feedback de Execução --- #
st.html('<h1 class="fonte_titulo_aula">Persistência de Dados e Feedback de Execução</h1>')
st.html('''<p class="fonte_texto">É fundamental destacar que loops de varredura que apenas consultam 
coordenadas sem preencher valores devem ser estruturados com cautela. Devido ao modelo de instanciação 
dinâmica de objetos do <span class='texto_python'>openpyxl</span>, navegar cegamente por um intervalo 
de coordenadas criará milhares de objetos de células vazias na memória RAM, o que pode impactar 
negativamente a performance de servidores e rotinas de big data.</p>''')
st.html('''<p class="fonte_texto">Depois de preencher todos os cabeçalhos e registros de dados na 
estrutura temporária criada na memória volátil, o processo de automação atinge o momento em que essas 
informações precisam ser transformadas em um arquivo físico permanente no sistema de armazenamento 
local.</p>''')
st.code("""# --- Salvar da memória para um arquivo Excel --- #
wb.save('vendas_iniciais.xlsx')
print('Arquivo salvo com sucesso!')""", line_numbers=True)
st.html('''<p class="fonte_texto">A chamada 
<span class='texto_python'>wb.save(</span><span class='variaveis'>'vendas_iniciais.xlsx'</span>
<span class='texto_python'>)</span> aciona os compiladores internos do 
<span class='texto_python'>openpyxl</span> para converter a hierarquia de objetos (
<i>Workbooks</i>, planilhas, células e valores) em arquivos XML individuais, compactando-os no formato 
unificado <span class='texto_python'>.xlsx</span>. O arquivo finalizado é gravado no diretório ativo 
onde o script Python está em execução.</p>''')
st.html('''<p class="fonte_texto">Por fim, a função 
<span class='funcoes_python'>print</span><span class='texto_python'>(</span>
<span class='variaveis'>'Arquivo salvo com sucesso!'</span><span class='texto_python'>)</span> é uma 
prática recomendada para a engenharia de software. Em sistemas automatizados que rodam em segundo plano 
ou em esteiras de integração contínua (CI/CD), a execução e a emissão de saídas de texto 
estruturadas no console funcionam como um log de sucesso primário, permitindo que operadores humanos 
ou ferramentas de monitoramento validem a integridade e o término adequado do fluxo sem a necessidade 
de inspecionar manualmente a existência do arquivo no sistema de arquivos.</p>''')
st.divider()

# --- Resumo --- #
st.html('<h1 class="fonte_titulo_aula">Resumo</h1>')
st.html('''<p class="fonte_texto">O desenvolvimento de scripts de automação com o 
<span class='texto_python'>openpyxl</span> exige o controle rigoroso do ciclo de vida dos dados, 
compreendendo as etapas que ocorrem desde a inicialização em memória até a gravação final em disco. 
Mapear cada comando utilizado no código para a sua respectiva ação física e propósito estratégico ajuda 
a criar fluxos muito mais organizados:</p>''')
st.html('<ul class="fonte_texto">'
        '<li><b>Instalação do Pacote:</b> Executa a instalação do pacote no terminal do sistema '
        'operacional por meio do gerenciador de pacotes <span class="texto_python">pip</span>. '
        'Estrategicamente, esse passo garante a integração dos decodificadores necessários para criar '
        'e processar a estrutura XML que compõe os arquivos de formato aberto.</li>'
        '<li><b>Instanciação do Livro de Trabalho:</b> Aloca e inicializa a classe estrutural do '
        'documento diretamente na memória RAM. Isso permite realizar operações de escrita e manipulação '
        'em alta velocidade sem gerar concorrência de gravação ou lentidão física no disco rígido do '
        'computador.</li>'
        '<li><b>Ativação da Planilha de Trabalho:</b> Retorna o objeto correspondente à aba que se '
        'encontra atualmente selecionada ou ativa para edição. Estrategicamente, fornece a referência de '
        'destino necessária para a aplicação começar a preencher dados na primeira guia.</li>'
        '<li><b>Renomeação da Aba:</b> Substitui o título automático atribuído pelo sistema (como o '
        'padrão <b>Sheet</b>) por um nome personalizado. O propósito prático dessa ação é conferir uma '
        'identidade clara e profissional ao relatório final apresentado ao usuário.</li>'
        '<li><b>Escrita Estática por Dicionário:</b> Identifica e acessa diretamente a célula '
        'especificada pela chave alfanumérica para preenchê-la com dados na memória. Essa abordagem é '
        'ideal para definir cabeçalhos ou campos isolados estáticos com máxima legibilidade de '
        'código.</li>'
        '<li><b>Escrita Dinâmica por Coordenadas:</b> Define a célula de destino utilizando coordenadas '
        'baseadas em números inteiros que representam a linha e a coluna. É um recurso estratégico '
        'fundamental para alimentar dados dinamicamente, permitindo integrar laços de repetição '
        'automáticos a estruturas tabulares.</li>'
        '<li><b>Persistência de Dados no Disco:</b> Transforma toda a hierarquia de objetos e dados '
        'voláteis em um arquivo físico estruturado e compactado no formato '
        '<span class="texto_python">.xlsx</span>. Isso consolida e salva permanentemente as informações '
        'criadas em seu disco de armazenamento local.</li>'
        '<li><b>Feedback de Execução:</b> Emite um indicador textual para a saída de dados padrão da '
        'linha de comando. Essa linha de código fornece rastreabilidade imediata, permitindo a verificação '
        'de logs em pipelines automáticos e garantindo que o programa concluiu todas as instruções de '
        'forma correta.</li>'
        '</ul>')
st.divider()

# --- Conclusão --- #
st.html('<h1 class="fonte_titulo_aula">Conclusão</h1>')
st.html('''<p class="fonte_texto">A utilização de scripts em Python para gerenciar tarefas rotineiras 
de manipulação de planilhas eletrônicas representa um ganho expressivo de performance e confiabilidade 
quando comparada a tarefas executadas de forma manual. Por intermédio da biblioteca 
<span class="texto_python">openpyxl</span>, dispomos de uma ponte entre o paradigma de 
programação orientada a objetos do Python e a representação tabular do formato OpenXML.</p>''')
st.html('''<p class="fonte_texto">O correto entendimento do gerenciamento de memória adotado pela 
biblioteca (que retém toda a estrutura em memória de forma volátil até a execução síncrona do método de 
salvamento) evita o gargalo de leitura e escrita em disco e permite a construção de pipelines ágeis. 
Adicionalmente, o domínio das duas técnicas de endereçamento de células (tanto a atribuição por chaves 
de dicionário quanto o preenchimento paramétrico por meio de coordenadas) torna a nossa vida mais fácil  
para projetar relatórios elegantes para analistas de negócios ou integrar 
complexos conjuntos de dados provenientes de bancos relacionais diretamente em planilhas prontas para 
o consumo corporativo.</p>''')
st.subheader('No mais é isso, nos vemos na próxima aula! Até lá, fiquem com Deus e fui!')
