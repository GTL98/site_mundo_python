# --- Importar as bibliotecas --- #
from PIL import Image
import streamlit as st

# --- Configurações iniciais da página --- #
st.set_page_config(
    page_title='Python para Excel - Aula 02',
    page_icon=Image.open('./assets/logo/logo.png'),
    layout='wide'
)

# --- Carregar o estilo das fontes --- #
with open('./assets/css/style.css', 'r') as css:
    st.html(f'<style>{css.read()}</style>')

# --- Colocar o título da aula --- #
st.html('<h1 class="fonte_titulo_aula">Aula 02: Múltiplas Abas, Iteração Inteligente e Tratamento de Dados!</h1>')

# --- Vídeo --- #
with st.expander('Se quiser acompanhar com o vídeo, acesse aqui! 👇'):
        st.video('https://youtu.be/Fkb8teblr3s')

# --- Código da aula --- #
st.subheader('Se quiser acessar o código completo da aula, clique [aqui](https://github.com/GTL98/canal_mundo_python/blob/main/Curso%20Completo%20de%20Python%20para%20Excel%3A%20Do%20Zero%20ao%20Especialista/Aula%2002/aula_02.ipynb)')
st.divider()

# --- Introdução --- #
st.subheader('E fala, devs! Tudo bem com vocês? Espero que sim!')
st.html('<h1 class="fonte_titulo_aula">Introdução</h1>')
st.html('''<p class="fonte_texto">A manipulação de planilhas eletrônicas é uma das tarefas mais 
recorrentes no cotidiano corporativo, servindo como base para relatórios financeiros, controle de 
inventários e auditorias de vendas. No entanto, a execução manual dessas atividades frequentemente 
introduz erros operacionais, além de consumir um tempo precioso das equipes técnicas. No ecossistema do 
Python, a biblioteca <span class='texto_python'>openpyxl</span> destaca-se como o padrão da indústria 
para a criação, edição e leitura de arquivos no formato binário do Microsoft Excel (
<span class='texto_python'>.xlsx</span>), permitindo autimatizemos fluxos de trabalho complexos com 
poucas linhas de código.</p>''')
st.html('''<p class="fonte_texto">Esta aula detalha como utilizar o Python para construir planilhas 
do Excel do zero, navegar de forma dinâmica entre múltiplas abas, processar registros horizontais 
utilizando técnicas de programação defensiva e auditar colunas verticalmente para extrair métricas de 
negócios confiáveis.</p>''')
st.subheader('Então sem mais delongas, bora para a aula!')
st.divider()

# --- Criação do Arquivo Excel --- #
st.html('<h1 class="fonte_titulo_aula">Criação do Arquivo Excel</h1>')
st.html('''<p class="fonte_texto">Para iniciar o desenvolvimento de um pipeline de dados com o Excel, o 
primeiro passo consiste em estruturar um arquivo físico contendo registros operacionais. No código a 
seguir, cria-se um arquivo de vendas contendo intencionalmente imperfeições estruturais (como linhas em 
branco e desvios de formatação de strings) a fim de simular as inconsistências comumente encontradas em 
bancos de dados reais e sistemas legados de ERP:</p>''')
st.code('''# --- Importar as bibliotecas --- #
import openpyxl
from datetime import datetime, date

# --- Criar o Workbook --- #
wb_origem = openpyxl.Workbook()

# --- Definir a aba ativa principal --- #
ws_sp = wb_origem.active
ws_sp.title = 'Faturamento_SP'

# --- Inserir dados com imperfeições --- #
ws_sp.append([])
ws_sp.append([datetime(2026, 3, 10), ' Notebook Pro ', 4500.00, 2])
ws_sp.append([])
ws_sp.append([])
ws_sp.append([datetime(2026, 3, 13), 'monitor gamer', 1899.99, 3])

# --- Adicionar dados à segundo aba/planilha --- #
ws_rj = wb_origem.create_sheet(title='Faturamento_RJ')
ws_rj.append([])
ws_rj.append([datetime(2026, 3, 10), 'Notebook Pro', 4600, 1])
ws_rj.append([])
ws_rj.append([])

# --- Salvar o arquivo Excel --- #
wb_origem.save('dados_vendas.xlsx')
print('Base de dados criada com sucesso!')''', line_numbers=True)
st.html('''<p class="fonte_texto">O processo de estruturação inicia-se com a importação do pacote 
<span class='texto_python'>openpyxl</span> e dos módulos <span class='texto_python'>datetime</span> 
e <span class='texto_python'>date</span>, necessários para lidar com dados temporais 
nativos do Python.</p>''')
st.html('''<p class="fonte_texto">A instanciação do objeto 
<span class='texto_python'>Workbook</span> por meio de 
<span class='texto_python'>openpyxl.Workbook()</span> cria uma representação conceitual de um arquivo 
Excel diretamente na memória RAM do computador. Isso evita o desgaste de escrita contínua em disco 
rígido durante o processo de montagem dos dados, garantindo maior desempenho na execução do script.</p>''')
st.html('''<p class="fonte_texto">Todo arquivo criado possui, por padrão, uma planilha inicial ativa. O 
comando <span class='texto_python'>ws_sp = wb_origem.active</span> obtém a referência dessa aba e, na 
linha subsequente, o atributo <span class='texto_python'>.title</span> é alterado para 
<span class='variaveis'>'Faturamento_SP'</span> para identificar que os dados pertencem à filial de 
São Paulo.</p>''')
st.html('''<p class="fonte_texto">A inserção de registros é feita utilizando o método 
<span class='texto_python'>.append()</span>, que aceita listas ou tuplas de dados e os insere na linha 
imediatamente inferior à última linha preenchida. A passagem de listas vazias (
<span class='texto_python'>[]</span>) simula a criação de linhas em branco acidentais. Essa técnica de 
indução de falhas é fundamental para validar a resiliência das rotinas de higienização de dados que serão 
detalhadas adiante. Além das linhas vazias, nota-se que o item 
<span class='variaveis'>' Notebook Pro '</span> foi cadastrado com espaçamentos a mais e 
<span class='variaveis'>'monitor gamer'</span> foi escrito inteiramente em letras minúsculas.</p>''')
st.html('''<p class="fonte_texto">Para incluir uma nova aba destinada à filial do Rio de Janeiro, 
utiliza-se o método <span class='texto_python'>.create_sheet(title=</span>
<span class='variaveis'>'Faturamento_RJ'</span><span class='texto_python'>)</span>. Após popular ambas 
as abas com registros mistos, a persistência física dos dados é realizada chamando 
<span class='texto_python'>wb_origem.save(</span>
<span class='variaveis'>'dados_vendas.xlsx'</span><span class='texto_python'>)</span>, gravando o 
binário no diretório de execução.</p>''')
colunas = st.columns(2)
with colunas[0]:
    st.subheader('Estrutura da Aba: Faturamento_SP')
    st.html('''<style type="text/css">
.tg  {border-collapse:collapse;border-spacing:0;}
.tg td{border-color:black;border-style:solid;border-width:1px;font-family:Arial, sans-serif;font-size:14px;
  overflow:hidden;padding:10px 5px;word-break:normal;}
.tg th{border-color:black;border-style:solid;border-width:1px;font-family:Arial, sans-serif;font-size:14px;
  font-weight:normal;overflow:hidden;padding:10px 5px;word-break:normal;}
.tg .tg-4rfe{border-color:#000000;color:#1F1F1F;font-family:Arial, Helvetica, sans-serif !important;font-size:16px;font-weight:bold;
  text-align:center;vertical-align:top}
.tg .tg-hsa2{font-family:Arial, Helvetica, sans-serif !important;font-size:16px;font-weight:bold;text-align:center;
  vertical-align:top}
.tg .tg-69a3{font-family:Arial, Helvetica, sans-serif !important;font-size:16px;text-align:center;vertical-align:top}
.tg .tg-lz2v{border-color:#000000;color:#1F1F1F;font-family:Arial, Helvetica, sans-serif !important;font-size:16px;
  text-align:center;vertical-align:top}
</style>
<table class="tg"><thead>
  <tr>
    <th class="tg-4rfe">Linha Física</th>
    <th class="tg-4rfe">Coluna A (Data)</th>
    <th class="tg-4rfe">Coluna B (Produto)</th>
    <th class="tg-hsa2">Coluna C (Preço)</th>
  </tr></thead>
<tbody>
  <tr>
    <td class="tg-lz2v">Linha 1</td>
    <td class="tg-lz2v">Vazia</td>
    <td class="tg-lz2v">Vazia</td>
    <td class="tg-69a3">Vazia</td>
  </tr>
  <tr>
    <td class="tg-lz2v">Linha 2 </td>
    <td class="tg-lz2v"><span class='texto_python'>2026-03-10 00:00:00</span></td>
    <td class="tg-lz2v"><span class='variaveis'>' Notebook Pro '</span></td>
    <td class="tg-69a3"><span class='numeros'>4500.00</span></td>
  </tr>
  <tr>
    <td class="tg-lz2v">Linha 3</td>
    <td class="tg-lz2v">Vazia</td>
    <td class="tg-lz2v">Vazia</td>
    <td class="tg-69a3">Vazia</td>
  </tr>
  <tr>
    <td class="tg-lz2v">Linha 4</td>
    <td class="tg-lz2v">Vazia</td>
    <td class="tg-lz2v">Vazia</td>
    <td class="tg-69a3">Vazia</td>
  </tr>
  <tr>
    <td class="tg-lz2v">Linha 5</td>
    <td class="tg-lz2v"><span class='texto_python'>2026-03-13 00:00:00</span></td>
    <td class="tg-lz2v"><span class='variaveis'>'monitor gamer'</span></td>
    <td class="tg-69a3"><span class='numeros'>1899.99</span></td>
  </tr>
</tbody></table>''')
with colunas[1]:
    st.subheader('Estrutura da Aba: Faturamento_RJ')
    st.html('''<style type="text/css">
.tg  {border-collapse:collapse;border-spacing:0;}
.tg td{border-color:black;border-style:solid;border-width:1px;font-family:Arial, sans-serif;font-size:14px;
  overflow:hidden;padding:10px 5px;word-break:normal;}
.tg th{border-color:black;border-style:solid;border-width:1px;font-family:Arial, sans-serif;font-size:14px;
  font-weight:normal;overflow:hidden;padding:10px 5px;word-break:normal;}
.tg .tg-4rfe{border-color:#000000;color:#1F1F1F;font-family:Arial, Helvetica, sans-serif !important;font-size:16px;font-weight:bold;
  text-align:center;vertical-align:top}
.tg .tg-hsa2{font-family:Arial, Helvetica, sans-serif !important;font-size:16px;font-weight:bold;text-align:center;
  vertical-align:top}
.tg .tg-69a3{font-family:Arial, Helvetica, sans-serif !important;font-size:16px;text-align:center;vertical-align:top}
.tg .tg-lz2v{border-color:#000000;color:#1F1F1F;font-family:Arial, Helvetica, sans-serif !important;font-size:16px;
  text-align:center;vertical-align:top}
</style>
<table class="tg"><thead>
  <tr>
    <th class="tg-4rfe">Linha Física</th>
    <th class="tg-4rfe">Coluna A (Data)</th>
    <th class="tg-4rfe">Coluna B (Produto)</th>
    <th class="tg-hsa2">Coluna C (Preço)</th>
  </tr></thead>
<tbody>
  <tr>
    <td class="tg-lz2v">Linha 1</td>
    <td class="tg-lz2v">Vazia</td>
    <td class="tg-lz2v">Vazia</td>
    <td class="tg-69a3">Vazia</td>
  </tr>
  <tr>
    <td class="tg-lz2v">Linha 2 </td>
    <td class="tg-lz2v"><span class='texto_python'>2026-03-10 00:00:00</span></td>
    <td class="tg-lz2v"><span class='variaveis'>'Notebook Pro'</span></td>
    <td class="tg-69a3"><span class='numeros'>4600</span></td>
  </tr>
  <tr>
    <td class="tg-lz2v">Linha 3</td>
    <td class="tg-lz2v">Vazia</td>
    <td class="tg-lz2v">Vazia</td>
    <td class="tg-69a3">Vazia</td>
  </tr>
  <tr>
    <td class="tg-lz2v">Linha 4</td>
    <td class="tg-lz2v">Vazia</td>
    <td class="tg-lz2v">Vazia</td>
    <td class="tg-69a3">Vazia</td>
  </tr>
</tbody>
</table>''')
st.divider()

# --- Navegação Dinâmica entre Múltiplas Abas --- #
st.html('<h1 class="fonte_titulo_aula">Navegação Dinâmica entre Múltiplas Abas</h1>')
st.html('''<p class="fonte_texto">O processamento manual de dados em planilhas frequentemente falha 
quando novas abas são inseridas ou quando seus nomes são modificados. O desenvolvimento de um código 
escalável exige que o carregamento do arquivo e a identificação das abas ocorram de forma totalmente 
dinâmica, permitindo que novas filiais ou períodos de faturamento sejam processados sem que haja 
necessidade de alterar o código-fonte:</p>''')
st.code('''# --- Caminho do arquivo --- #
caminho_arquivo = 'dados_vendas.xlsx'

# --- Carregar o arquivo --- #
wb_leitura = openpyxl.load_workbook(
    filename=caminho_arquivo,  # arquivo a ser carregado
    data_only=True  # mostra somente os dados, se tiver uma fórmula; mostra o resultado, não a fórmula em si
)

# --- Saber quais abas estão ativas no Workbook --- #
abas_disponiveis = wb_leitura.sheetnames
print(f'Abas identificadas de forma dinâmica no arquivo: {abas_disponiveis}')''', line_numbers=True)
st.html('''<p class="fonte_texto">A função <span class='texto_python'>load_workbook</span> do pacote 
<span class='texto_python'>openpyxl</span> é a encarregada de ler o arquivo 
<span class='texto_python'>.xlsx</span> do disco e transformá-lo de volta em um objeto manipulável em 
memória Python. O grande destaque analítico nesta etapa reside na configuração do parâmetro 
<span class='texto_python'>data_only=</span><span class='palavras_reservadas'>True</span>.</p>''')
st.html('''<p class="fonte_texto">No Microsoft Excel, uma célula que possui uma fórmula dinâmica 
(como <span class='texto_python'>=SOMA(C2:C5)</span>) armazena internamente tanto a estrutura de texto da 
fórmula quanto o último valor numérico calculado por ela. Se o arquivo for aberto com 
<span class='texto_python'>data_only=</span><span class='palavras_reservadas'>False</span> (padrão da 
biblioteca), o Python receberá a string contendo a fórmula literal. Caso se tente realizar operações 
matemáticas diretamente com essa string em nosso código, o interpretador gerará erros de tipo. Ao 
definirmos <span class='texto_python'>data_only=</span><span class='palavras_reservadas'>True</span>, 
instruímos o leitor da biblioteca a ignorar as fórmulas textuais e extrair unicamente os valores 
numéricos brutos calculados.</p>''')
st.html('''<p class="fonte_texto">Para mapear a estrutura interna do arquivo carregado de forma 
automática, acessa-se a propriedade <span class='texto_python'>.sheetnames</span>. Esse atributo 
examina os metadados do arquivo Excel e retorna uma lista contendo os nomes de todas as abas 
atualmente configuradas no arquivo. A partir dessa extração, o script ganha adaptabilidade para 
iterar por todas as abas/planilhas disponíveis.</p>''')
st.divider()

# --- Iteração Eficiente via iter_rows e Tratamento de Tipos --- #
st.html('<h1 class="fonte_titulo_aula">Iteração Eficiente via <span class="texto_python">iter_rows</span> '
        'e Tratamento de Tipos</h1>')
st.html('''<p class="fonte_texto">A varredura horizontal foca no processamento sequencial de cada 
registro completo (ou seja, cada linha que agrupa atributos de uma transação específica). Para tanto, o 
gerador <span class='texto_python'>iter_rows()</span> é a ferramenta mais indicada por permitir o 
estabelecimento de limites de busca, ignorar cabeçalhos desnecessários e extrair os valores diretamente 
das células, reduzindo significativamente o consumo de memória RAM durante a execução:</p>''')
st.code(r'''# --- Iteração por linhas --- #
limite_valor_premium = 2000
faturamento_premium_total = 0

print('--- INICIANDO EXTRAÇÃO DE DADOS POR REGISTRO (ITER_ROWS) ---')
for aba in abas_disponiveis:
    aba_atual = wb_leitura[aba]

    # --- O iter_rows ignora a primeira linha e restringe a busca em 4 colunas --- #
    for linha in aba_atual.iter_rows(min_row=2, max_col=4, values_only=True):
        data_bruta, produto_bruto, preco_bruto, quantidade_bruta = linha

        # --- Validação defensiva contra linhas vazias ou nulas --- #
        if data_bruta is None or produto_bruto is None or preco_bruto is None:
            continue

        # --- Tratamento de strings: eliminação de espaçamentos e padronização textual --- #
        produto_limpo = str(produto_bruto).strip().title()

        # --- Tratamento de floats: assegurar coerência numérica para operações matemáticas --- #
        preco_numerico = float(preco_bruto)

        # --- Tratamento de datetimes: conversão para visualização no padrão brasileiro --- #
        if isinstance(data_bruta, (datetime, date)):
            data_formatada = data_bruta.strftime('%d/%m/%Y')
        else:
            data_formatada = str(data_bruta)

        # --- Extração de inteligência: filtro de transações de alto valor (Premium) --- #
        if preco_numerico >= limite_valor_premium:
            subtotal = preco_numerico * int(quantidade_bruta)
            faturamento_premium_total += subtotal
            print(f'[{aba}] Registro Processado: Data {data_formatada} | '
f'Item: {produto_limpo} | Valor: R$ {preco_numerico:,.2f} | '
f'Qtd: {quantidade_bruta} | Subtotal: R$ {subtotal:,.2f}')

print('\nInteligência consolidada:')
print(f'Faturamento total acumulado de itens premium (>= R$ 2.000,00): R$ {faturamento_premium_total:,.2f}')''', line_numbers=True)
st.html('''<p class="fonte_texto">A estrutura de repetição externa percorre as abas descobertas 
dinamicamente (<span class='palavras_reservadas'>for</span> <span class='texto_python'>aba</span> 
<span class='palavras_reservadas'>in</span> <span class='texto_python'>abas_disponiveis</span>) e carrega 
o objeto de planilha ativa utilizando o nome retornado: 
<span class='texto_python'>aba_atual = wb_leitura[aba]</span>.</p>''')
st.html('''<p class="fonte_texto">O laço interno utiliza o método 
<span class='texto_python'>iter_rows()</span>. O método é parametrizado da seguinte forma:</p>''')
st.html('''<ul class="fonte_texto">
        <li><span class='texto_python'>min_row=</span><span class='numeros'>2</span>: Instrui o 
        interpretador a iniciar a varredura a partir da segunda linha física da planilha. Isso evita que 
        a linha de cabeçalhos ou a primeira linha vazia do nosso conjunto de testes causem erros de 
        descompactação.</li>
        <li><span class='texto_python'>max_col=</span><span class='numeros'>4</span>: Limita a leitura às 
        primeiras quatro colunas (da coluna A até a coluna D). Essa restrição impede que o Python 
        continue lendo infinitamente colunas vazias à direita dos dados reais, otimizando o tempo de 
        processamento.</li>
        <li><span class='texto_python'>values_only=</span><span class='palavras_reservadas'>True</span>: 
        Por padrão, o <span class='texto_python'>openpyxl</span> retorna objetos do tipo 
        <span class='texto_python'>Cell</span>. Para acessarmos o valor da célula, precisaríamos chamar o 
        atributo <span class='texto_python'>.value</span> em cada um deles. Ao definir este parâmetro 
        como verdadeiro, o gerador retorna diretamente uma tupla contendo os valores puros armazenados 
        nas células, o que simplifica a leitura e agiliza o desempacotamento.</li>
        </ul>''')
st.html('''<p class="fonte_texto">A descompactação de tuplas ocorre na instrução 
<span class='texto_python'>data_bruta</span>, <span class='texto_python'>produto_bruto, preco_bruto, 
quantidade_bruta = linha</span>. Caso a linha atual esteja vazia, ela retornará uma tupla repleta de 
valores nulos (<span class='texto_python'>(</span><span class='palavras_reservadas'>None</span>
<span class='texto_python'>,</span> <span class='palavras_reservadas'>None</span>
<span class='texto_python'>,</span> <span class='palavras_reservadas'>None</span>
<span class='texto_python'>,</span> <span class='palavras_reservadas'>None</span>
<span class='texto_python'>)</span>). A verificação de segurança 
<span class='palavras_reservadas'>if</span> <span class='texto_python'>data_bruta</span> 
<span class='palavras_reservadas'>is None or</span><span class='texto_python'>...:</span> 
<span class='palavras_reservadas'>continue</span> constitui uma importante técnica de programação 
defensiva. Ao detectar qualquer valor nulo nos atributos principais do registro de vendas, o 
interpretador interrompe o fluxo da iteração corrente com a instrução 
<span class='palavras_reservadas'>continue</span> e avança para a próxima linha sem executar operações 
que resultariam em falhas graves.</p>''')
st.html('''<p class="fonte_texto">Uma vez validado que o registro contém dados consistentes, inicia-se 
a etapa de sanitização e conversão de tipos:</p>''')
st.html('''<ol type=1 class="fonte_texto">
        <li><b>Strings:</b> Os nomes de produtos vindos do Excel frequentemente contêm espaçamentos 
        adicionais digitados acidentalmente. O comando 
        <span class='funcoes_python'>str</span>
        <span class='texto_python'>(produto_bruto).strip().title()</span> primeiro remove esses 
        espaçamentos nas extremidades da palavra através do método 
        <span class='texto_python'>.strip()</span> e, na sequência, aplica 
        <span class='texto_python'>.title()</span> para capitalizar a primeira letra de cada palavra. 
        Isso garante que <span class='variaveis'>' Notebook Pro '</span> seja limpo e padronizado como 
        <span class='variaveis'>'Notebook Pro'</span> e que 
        <span class='variaveis'>'monitor gamer'</span> passe a figurar de forma elegante como 
        <span class='texto_python'>'Monitor Gamer'</span>.</li>
        <li><b>Números:</b> O preço bruto é convertido explicitamente para um valor de ponto flutuante 
        (<span class='funcoes_python'>float</span><span class='texto_python'>(preco_bruto)</span>), 
        assegurando que operações matemáticas possam ser realizadas sem conflito de tipos de dados.</li>
        <li><b>Datas:</b> O Excel gerencia datas como objetos do tipo data e hora do próprio sistema 
        operacional ou do Python. O bloco condicional 
        <span class='funcoes_python'>isinstance</span>
        <span class='texto_python'>(data_bruta, (datetime, date)</span>) valida se o tipo corresponde a 
        esses objetos temporais. Em caso positivo, o método 
        <span class='texto_python'>.strftime(</span><span class='variaveis'>'%d/%m/%Y'</span>
        <span class='texto_python'>)</span> traduz a representação de máquina para a notação visual 
        brasileira de dia/mês/ano. Caso o valor seja apenas uma string, ela é convertida diretamente 
        pelo construtor de string nativo.</li>
        </ol>''')
st.html('''<p class="fonte_texto">O fechamento da rotina aplica uma regra de negócio simples: se o preço 
do produto limpo atingir o patamar configurado como premium (R$2.000,00), calcula-se o subtotal através 
da fórmula:</p>''')
st.latex(r'\text{Subtotal} = \text{Preço Numérico} \times \text{Quantidade}')
st.html('''<p class="fonte_texto">Este subtotal é acumulado na variável 
<span class='texto_python'>faturamento_premium_total</span> e os dados processados são exibidos de forma 
organizada para fins de auditoria.</p>''')
st.divider()

# --- Inspeção Vertical e Auditoria via iter_cols --- #
st.html('<h1 class="fonte_titulo_aula">Inspeção Vertical e Auditoria via '
        '<span class="texto_python">iter_cols</span></h1>')
st.html('''<p class="fonte_texto">Enquanto a varredura por linhas foca em transações individuais, a 
inspeção vertical por colunas é ideal para analisar métricas financeiras consolidadas ao longo de toda 
uma categoria de dados (como o preço médio praticado por cada filial). A utilização de 
<span class="texto_python">iter_cols()</span> possibilita isolar uma única coluna de interesse no 
arquivo do Excel, evitando o carregamento desnecessário de outras colunas:</p>''')
st.code('''# --- Inspeção vertical --- #
print('--- INICIADO EXTRAÇÃO DE COLUNAS COM ITER_COLS ---')
for aba in abas_disponiveis:
    aba_atual = wb_leitura[aba]

    # --- O método iter_cols isola apenas a terceira coluna, pulando o cabeçalho --- #
    for coluna_precos in aba_atual.iter_cols(min_row=2, min_col=3, max_col=3, values_only=True):
        # --- Filtrar elementos nulos para evitar insconsistências analíticas --- #
        lista_precos = [float(p) for p in coluna_precos if p is not None]

        # --- Operações estatísticas --- #
        if lista_precos:
            ticket_medio_filial = sum(lista_precos) / len(lista_precos)
            maior_preco_filial = max(lista_precos)
        else:
            ticket_medio_filial = 0
            maior_preco_filial = 0

        print(f'Auditoria da filial: [{aba}] | Ticket médio de venda: R$ {ticket_medio_filial:,.2f} | '
f'Maior preço identificado: R$ {maior_preco_filial:,.2f}')''', line_numbers=True)
st.html('''<p class="fonte_texto">A função <span class="texto_python">iter_cols()</span> é parametrizada 
com restrições rígidas para otimizar os recursos do interpretador:</p>''')
st.html('''<ul class="fonte_texto">
        <li><span class="texto_python">min_row=</span><span class="numeros">2</span>: Mantém a lógica de pular 
        o cabeçalho descritivo que se encontra na primeira linha física da coluna de preços.</li>
        <li><span class="texto_python">min_col=</span><span class="numeros">3</span> e 
        <span class="texto_python">max_col=</span><span class="numeros">3</span>: Delimita que a inspeção 
        deve se concentrar única e exclusivamente na coluna de índice 3 (coluna C, que armazena os 
        valores monetários de venda). Ao focar na coluna de interesse de forma isolada, evita-se a 
        leitura de dados de texto e quantidades, o que acelera o processamento em bases extensas.</li>
        <li><span class="texto_python">values_only=</span><span class="palavras_reservadas">True</span>: 
        Garante que o gerador forneça diretamente a tupla contendo os valores dos preços contidos na 
        coluna.</li>
        </ul>''')
st.html('''<p class="fonte_texto">Como as planilhas criadas de forma simulada contêm células vazias no 
final de cada bloco de transações, a tupla retornada por <span class="texto_python">iter_cols()</span> 
incluirá objetos do tipo <span class="palavras_reservadas">None</span>. Se tentássemos aplicar operações 
estatísticas diretamente sobre essa coleção, o Python geraria exceções insolúveis de matemática 
financeira.</p>''')
st.html('''<p class="fonte_texto">A resolução deste problema é alcançada de forma limpa e concisa por 
meio de uma <i>list comprehension</i>: <span class="texto_python">[</span>
<span class="funcoes_python">float</span><span class="texto_python">(p)</span> 
<span class="palavras_reservadas">for</span> <span class="texto_python">p</span> 
<span class="palavras_reservadas">in</span> <span class="texto_python">coluna_precos</span> 
<span class="palavras_reservadas">if</span> <span class="texto_python">p</span> 
<span class="palavras_reservadas">is not None</span><span class="texto_python">]</span>. Esta linha 
atua como um filtro ativo, convertendo e adicionando à <span class="texto_python">lista_precos</span> 
somente os valores válidos e descartando todas as entradas nulas antes do cálculo estatístico.</p>''')
st.html('''<p class="fonte_texto">A determinação do ticket médio e do preço máximo utiliza métodos 
integrados de forma altamente eficiente:</p>''')
st.html('''<ul class="fonte_texto">
        <li><b>Ticket Médio (TM):</b> Calculado dividindo a soma de todos os preços filtrados pelo total de 
        elementos válidos encontrados na lista, conforme expressa a fórmula matemática:</li>
        </ul>''')
st.latex(r'\text{TM} = \frac{\sum_{i=1}^{n} P_i}{n}')
st.html('''<p class="fonte_texto">Onde <i>Pi</i> representa cada preço e <i>n</i> é a quantidade total 
de registros numéricos válidos. O bloco condicional <span class="palavras_reservadas">if</span> 
<span class="texto_python">lista_precos:</span> garante de forma defensiva que a divisão por zero não 
ocorra em planilhas que eventualmente estejam vazias ou sem valores válidos.</p>''')
st.html('''<ul class="fonte_texto">
        <li><b>Maior Preço:</b> A função nativa <span class="funcoes_python">max</span>
        <span class="texto_python">(lista_precos)</span> vasculha a lista resultante e retorna 
        instantaneamente o valor máximo de venda encontrado para aquela filial.</li>
        </ul>''')
st.html('''<p class="fonte_texto">Para compreender as distinções técnicas essenciais e a correta 
aplicação de cada um dos métodos iterativos estudados nesta aula, apresenta-se a tabela de 
comparação estrutural abaixo:</p>''')
st.html('''<style type="text/css">
.tg  {border-collapse:collapse;border-spacing:0;}
.tg td{border-color:black;border-style:solid;border-width:1px;font-family:Arial, sans-serif;font-size:14px;
  overflow:hidden;padding:10px 5px;word-break:normal;}
.tg th{border-color:black;border-style:solid;border-width:1px;font-family:Arial, sans-serif;font-size:14px;
  font-weight:normal;overflow:hidden;padding:10px 5px;word-break:normal;}
.tg .tg-4rfe{border-color:#000000;color:#1F1F1F;font-family:Arial, Helvetica, sans-serif !important;font-size:16px;font-weight:bold;
  text-align:center;vertical-align:top}
.tg .tg-lz2v{border-color:#000000;color:#1F1F1F;font-family:Arial, Helvetica, sans-serif !important;font-size:16px;
  text-align:center;vertical-align:top}
</style>
<table class="tg"><thead>
  <tr>
    <th class="tg-4rfe">Atributo de Comparação</th>
    <th class="tg-4rfe">Método <span class="texto_python">iter_rows</span></th>
    <th class="tg-4rfe">Método <span class="texto_python">iter_cols</span></th>
  </tr></thead>
<tbody>
  <tr>
    <td class="tg-lz2v">Direção de varredura</td>
    <td class="tg-lz2v">Horizontal (linha por linha)</td>
    <td class="tg-lz2v">Vertical (coluna por coluna)</td>
  </tr>
  <tr>
    <td class="tg-lz2v">Caso de uso principal</td>
    <td class="tg-lz2v">Processar registros de transações<br>completos (dados que dependem<br>uns dos outros ao longo da mesma linha)</td>
    <td class="tg-lz2v">Analisar métricas estatísticas<br>consolidadas e realizar auditorias de<br>valores de um único atributo (como<br>colunas de preço)</td>
  </tr>
  <tr>
    <td class="tg-lz2v">Formato de retorno</td>
    <td class="tg-lz2v">Tuplas contendo os valores de<br>múltiplas células em sequência horizontal</td>
    <td class="tg-lz2v">Tuplas contento os valores de<br>múltiplas células em sequência vertical<br>de um coluna isolada</td>
  </tr>
  <tr>
    <td class="tg-lz2v">Otimização de memória</td>
    <td class="tg-lz2v">Alta eficiência para filtragens e<br>checagens estruturais de<br>registros completos</td>
    <td class="tg-lz2v">Excelente performance ao limitar a<br>busca a colunas específicas, evitando<br>ler dados adjacentes irrelevantes</td>
  </tr>
</tbody></table>''')
st.divider()

# --- Resumo --- #
st.html('<h1 class="fonte_titulo_aula">Resumo</h1>')
st.html('''<p class="fonte_texto">Esta aula apresentou os aspectos teóricos e práticos que fundamentam 
o desenvolvimento profissional de pipelines de automação para planilhas eletrônicas utilizando a 
biblioteca <span class="texto_python">openpyxl</span> em Python.</p>''')
st.html('''<p class="fonte_texto">Inicialmente, vimos a criação estruturada de arquivos do Excel (
<span class="texto_python">.xlsx</span>), abordando a hierarquia que governa a relação entre pastas de 
trabalho (<i>Workbook</i>), abas (<i>Worksheet</i>) e células (<i>Cell</i>). Mostramos como povoar 
esses arquivos programaticamente e salvá-los de forma consistente no ambiente de trabalho.</p>''')
st.html('''<p class="fonte_texto">Dando sequência ao fluxo de processamento de dados, abordamos o 
carregamento e a exploração de arquivos por meio de rotinas de mapeamento dinâmico de abas com o 
atributo <span class="texto_python">.sheetnames</span>. Detalhamos a extrema importância do parâmetro 
<span class="texto_python">data_only=</span><span class="palavras_reservadas">True</span> como o 
elemento habilitador da captura direta de resultados gerados por fórmulas complexas do Excel.</p>''')
st.html('''<p class="fonte_texto">Por fim, vimos a aplicação prática dos geradores de iteração 
horizontal e vertical:</p>''')
st.html('''<ul class="fonte_texto">
        <li>O método <span class="texto_python">iter_rows()</span> foi utilizado para inspecionar 
        registros de forma horizontal, permitindo pular cabeçalhos e trabalhar diretamente com tuplas de 
        valores brutos. Nesse escopo, aprendemos técnicas de programação defensiva para neutralizar 
        valores vazios (<span class="palavras_reservadas">None</span>) e sanitizar strings de produtos 
        utilizando os métodos <span class="texto_python">.strip()</span> e 
        <span class="texto_python">.title()</span>.</li>
        <li>O método <span class="texto_python">iter_cols()</span> foi introduzido para a execução de 
        inspeções verticais estruturadas sobre colunas específicas. Aprendemos como isolar a coluna de 
        preços, filtrar valores nulos utilizando <i>list comprehensions</i> para eliminar erros em 
        operações estatísticas e, por meio desse conjunto filtrado, computar de forma precisa o ticket 
        médio de vendas e o maior preço praticado por cada uma das filiais auditadas.</li>
        </ul>''')
st.divider()

# --- Conclusão --- #
st.html('<h1 class="fonte_titulo_aula">Conclusão</h1>')
st.html('''<p class="fonte_texto">A automação de planilhas com Python e 
<span class="texto_python">openpyxl</span> consolida-se como um divisor de águas na atuação de 
analistas e engenheiros de dados. Enquanto processos manuais de formatação, limpeza e extração de 
relatórios são inerentemente propensos a erros e difíceis de escalar, a adoção de rotinas programadas 
assegura integridade absoluta dos dados de ponta a ponta e escalabilidade sem precedentes.</p>''')
st.html('''<p class="fonte_texto">Ao utilizar de forma combinada o poder da programação defensiva e as 
técnicas de iteração inteligente oferecidas por 
<span class="texto_python">iter_rows</span> e <span class="texto_python">iter_cols</span>, podemos 
não apenas minimizam gargalos de memória em ambientes corporativos, como também criam scripts 
altamente adaptáveis a variações na quantidade de abas ou no volume de dados. Assim, rotinas complexas 
de cálculo de comissões, faturamentos e auditorias podem ser reexecutadas instantaneamente a cada 
período, nos liberando para focar em tarefas de cunho analítico e estratégico.</p>''')
st.subheader('No mais é isso, nos vemos na próxima aula! Até lá, fiquem com Deus e fui!')