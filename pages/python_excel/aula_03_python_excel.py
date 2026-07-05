# --- Importar as bibliotecas --- #
from PIL import Image
import streamlit as st
from carregar_css import carregar_css

# --- Configurações iniciais da página --- #
st.set_page_config(
    page_title='Python para Excel - Aula 03',
    page_icon=Image.open('./assets/logo/logo.png'),
    layout='wide'
)

# --- Carregar o CSS --- #
carregar_css()

# --- Colocar o título da aula --- #
st.html('<h1 class="fonte_titulo_aula">Aula 03: Estilização Profissional – Cores, Fontes, Bordas e Formatação!</h1>')

# --- Vídeo --- #
with st.expander('Se quiser acompanhar com o vídeo, acesse aqui! 👇'):
        st.video('https://youtu.be/Pop-4LrQAx4')

# --- Código da aula --- #
st.subheader('Se quiser acessar o código completo da aula, clique [aqui](https://github.com/GTL98/canal_mundo_python/blob/main/Curso%20Completo%20de%20Python%20para%20Excel%3A%20Do%20Zero%20ao%20Especialista/Aula%2003/aula_03.ipynb)')
st.divider()

# --- Introdução --- #
st.subheader('E fala, devs! Tudo bem com vocês? Espero que sim!')
st.html('<h1 class="fonte_titulo_aula">Introdução</h1>')
st.html('''<p class="fonte_texto">Se você está cansado de gerar relatórios no Excel de forma manual ou de 
entregar planilhas com aquela cara cinza e sem graça que ninguém tem paciência de analisar, você está no 
lugar certo. Hoje daremos o pontapé inicial em uma jornada incrível para automatizar e transformar 
planilhas comuns em relatórios visuais extremamente profissionais utilizando o poder do Python!</p>''')
st.html('''<p class="fonte_texto">A ideia aqui não é apenas jogar os dados na planilha, mas sim construir 
uma automação robusta onde o Python cuida de todo o trabalho duro e chato para você. Imagine criar uma 
estrutura, popular com informações, formatar moedas, ajustar larguras e aplicar um design elegante com 
apenas um clique. É exatamente isso que construiremos passo a passo, começando do absoluto zero, para 
que você possa aplicar essa lógica em qualquer relatório do seu dia a dia.</p>''')
st.subheader('Então sem mais delongas, bora para a aula!')
st.divider()

# --- Construção da base de dados e estrutura inicial --- #
st.html('<h1 class="fonte_titulo_aula">Construção da base de dados e estrutura inicial</h1>')
st.html('''<p class="fonte_texto">Para fazer essa mágica acontecer, a biblioteca que nos acompanhará é 
a <span class="texto_python">openpyxl</span>, uma das ferramentas mais poderosas do ecossistema Python 
para manipulação de arquivos 
<span class="texto_python">.xlsx</span>. Antes de colocar a mão na massa, é fundamental entender que o 
<span class="texto_python">openpyxl</span> enxerga um arquivo do Excel sob a ótica de dois conceitos 
principais: o <b>Workbook</b> (que é o arquivo de trabalho, a pasta de arquivos em si) e a 
<b>Worksheet</b> (que é a planilha ou aba específica dentro desse arquivo).</p>''')
st.html('''<p class="fonte_texto">Nesta primeira etapa, o nosso foco total está em erguer os pilares do 
nosso projeto. Vamos criar o arquivo na memória, definir o nome da nossa aba de faturamento, inserir um 
título principal de destaque, mapear os cabeçalhos das colunas e preencher as linhas com o nosso conjunto 
de dados de vendas. Fazer essa separação e estruturação limpa dos dados logo no início é uma excelente 
prática de programação, pois garante que a nossa base esteja sólida antes de partirmos para a camada de 
estilização visual:</p>''')
st.code('''# --- Importar os módulos --- #
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# --- Inicialização do Workbook --- #
wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Faturamento'

# --- Texto de título na célula --- #
ws['A1'] = 'Relatório de Faturamento Mensal'

# --- Definição e inserção dos cabeçalhos --- #
cabecalhos = ['ID', 'Produto', 'Quantidade', 'Receita']
for col_num, cabecalho in enumerate(cabecalhos, 1):
    ws.cell(row=2, column=col_num, value=cabecalho)

# --- Conjunto de dados --- #
dados = [
    [101, 'Notebook Pro', 5, 25_000.00],
    [102, 'Monitor Ultrawide', 12, 18_000.50],
    [103, 'Teclado Mecânico', 25, 3750.00],
    [104, 'Mouse Ergonômico', 30, 4500.25]
]

# --- Colocar os dados nas células correspondentes --- #
for indice_linha, dado_linha in enumerate(dados, 3):
    for indice_coluna, dado in enumerate(dado_linha, 1):
        ws.cell(row=indice_linha, column=indice_coluna, value=dado)

wb.save('arquivo_01.xlsx')''', line_numbers=True)
st.html('''<p class="fonte_texto">Vamos destrinchar cada bloco desse código para que você entenda 
exatamente o que está acontecendo por baixo dos panos.</p>''')
st.html('''<p class="fonte_texto">No topo do código, fazemos a importação do 
<span class="texto_python">openpyxl</span>. Note que também já estamos trazendo módulos bem específicos de 
dentro dele, como <span class="texto_python">get_column_letter</span> (que nos ajudará a converter números 
em letras de colunas mais adiante) e várias classes de estilização como 
<span class="texto_python">Font</span>, <span class="texto_python">PatternFill</span>, 
<span class="texto_python">Alignment</span>, <span class="texto_python">Border</span> e 
<span class="texto_python">Side</span>. Mesmo que a gente vá usar a pintura e as fontes nas próximas 
etapas, já é uma ótima prática deixar tudo importado e preparado para evitar poluição visual no código 
depois.</p>''')
st.html('''<p class="fonte_texto">Com o comando 
<span class="texto_python">wb = openpyxl.Workbook()</span>, nós efetivamente criamos um arquivo de Excel 
novinho em folha na memória do computador. Quando um arquivo novo é criado, o Excel automaticamente gera 
uma aba padrão. Nós capturamos essa aba ativa com <span class="texto_python">ws = wb.active</span>. 
Em seguida, para deixar o projeto organizado e com cara de relatório corporativo, alteramos o nome dessa 
aba de "Sheet" para "Faturamento" usando 
<span class="texto_python">ws.title = </span><span class="variaveis">'Faturamento'</span>.</p>''')
st.html('''<p class="fonte_texto">A linha <span class="texto_python">ws[</span>
<span class="variaveis">'A1'</span><span class="texto_python">] = </span>
<span class="variaveis">'Relatório de Faturamento Mensal'</span> faz exatamente o que parece: ela vai 
direto na célula <b>A1</b> (coluna A, linha 1) da nossa planilha e escreve o texto do nosso título. Por 
enquanto ele parecerá apenas um texto comum, mas essa célula será a base para a nossa futura 
mesclagem.</p>''')
st.html('''<p class="fonte_texto">Criamos uma lista simples chamada 
<span class="texto_python">cabecalhos</span> contendo os títulos das nossas colunas: ID, Produto, 
Quantidade e Receita. Para inseri-los na planilha, utilizamos um loop 
<span class="palavras_reservadas">for</span> combinado com a função 
<span class="funcoes_python">enumerate</span><span class="texto_python">()</span>.</p>''')
st.html('''<p class="fonte_texto">O <span class="funcoes_python">enumerate</span>
<span class="texto_python">(cabecalhos, </span><span class="numeros">1</span>
<span class="texto_python">)</span> faz algo fantástico. Ele percorre a lista e nos devolve 
duas coisas a cada rodada: o nome do cabeçalho e o número da posição dele. Passamos o parâmetro 
<span class="numeros">1</span> para indicar que a contagem deve começar em 1, e não em 0 (que é o padrão 
do Python). Por que fazemos isso? Porque o Excel começa a contar suas colunas a partir do número 1 
(Coluna 1 = A, Coluna 2 = B, etc.).</p>''')
st.html('''<p class="fonte_texto">Dentro do loop, usamos 
<span class="texto_python">ws.cell(row=</span><span class="numeros">2</span>
<span class="texto_python">, column=col_num, value=cabecalho)</span>. Isso diz ao Python: 
<i>"Vá na linha 2, mude de coluna a cada rodada do loop (1, 2, 3, 4) e coloque o texto correspondente"</i>. 
Usamos a linha 2 porque a linha 1 já está guardada para o nosso título principal.</p>''')
st.html('''<p class="fonte_texto">A variável <span class="texto_python">dados</span> armazena uma lista 
de listas. Cada lista interna representa uma linha horizontal completa da nossa futura tabela, contendo o 
ID do produto, o nome dele, a quantidade vendida e o valor total de receita. Um detalhe muito legal de 
sintaxe Python aqui são os sublinhados nos números (ex: <span class="numeros">25_000.00</span>). O Python 
ignora esses underscores, mas para nós, programadores, fica infinitamente mais fácil ler que o número se 
trata de vinte e cinco mil.</p>''')
st.html('''<p class="fonte_texto">Para descarregar essa matriz de dados dentro do Excel, usamos dois 
loops aninhados (um dentro do outro), que funcionam como uma varredura de grade:</p>''')
st.html('''<ul class="fonte_texto">
<li>O primeiro loop 
<span class="palavras_reservadas">for </span><span class="texto_python">indice_linha, dado_linha </span>
<span class="palavras_reservadas">in </span><span class="funcoes_python">enumerate</span>
<span class="texto_python">(dados, </span><span class="numeros">3</span>
<span class="texto_python">)</span> percorre cada linha da nossa matriz de dados. Repare que passamos o 
número <span class="numeros">3</span> no <span class="texto_python">enumerate</span>. Isso é crucial 
porque a linha 1 é o título e a linha 2 são os cabeçalhos. Portanto, nossos dados reais precisam começar 
obrigatoriamente na linha 3.</li>
<li>O segundo loop <span class="palavras_reservadas">for </span><span class="texto_python">indice_coluna, 
dado </span>
<span class="palavras_reservadas">in </span><span class="funcoes_python">enumerate</span>
<span class="texto_python">(dado_linha, </span><span class="numeros">1</span>
<span class="texto_python">)</span> entra dentro da linha atual e percorre cada item individual (coluna 
por coluna), começando da coluna 1 (coluna A).</li>
<li>A mágica acontece em <span class="texto_python">ws.cell(row=indice_linha, column=indice_coluna, 
value=dado)</span>. O Python navegará dinamicamente (Linha 3 Coluna 1, Linha 3 Coluna 2...) e 
soltando o valor exato na célula correspondente.</li>
</ul>''')
st.html('''<p class="fonte_texto">Por fim, executamos 
<span class="texto_python">wb.save(</span><span class="variaveis">'arquivo_01.xlsx'</span>
<span class="texto_python">)</span>. Esse comando pega tudo o que estruturamos na memória do computador 
e joga para um arquivo físico real no seu diretório. Ao abrir esse arquivo, você verá todos os dados 
perfeitamente organizados em linhas e colunas.</p>''')
st.divider()

# --- Mesclagem e comunicação visual de títulos --- #
st.html('<h1 class="fonte_titulo_aula">Mesclagem e comunicação visual de títulos</h1>')
st.html('''<p class="fonte_texto">Com a nossa base de dados perfeitamente construída e estruturada, é hora 
de começar a parte mais divertida: dar vida e cor à nossa planilha! Afinal, um relatório de respeito 
precisa ter uma identidade visual que prenda a atenção de quem está lendo. Nesta segunda parte, vamos 
focar exclusivamente no nosso título principal, transformando aquele texto simples que jogamos na célula 
A1 em um cabeçalho imponente e com aspecto corporativo.</p>''')
st.html('''<p class="fonte_texto">Para alcançarmos esse resultado visual, entraremos na parte de 
estilização da biblioteca <span class="texto_python">openpyxl</span>. Teoricamente, a estilização no 
Excel via Python funciona como se estivéssemos operando as opções da barra de ferramentas do próprio 
Excel, mas usando código. Precisaremos dominar quatro conceitos básicos aqui: a <b>mesclagem de 
células</b> (para criar um grande bloco contínuo de espaço), a formatação de <b>fonte</b> (tamanho, cor e 
estilo), o <b>preenchimento de fundo</b> (cores sólidas usando sistema hexadecimal) e o 
<b>alinhamento</b>.</p>''')
st.html('''<p class="fonte_texto">Um detalhe técnico super importante sobre o funcionamento do 
<span class="texto_python">openpyxl</span> que abordaremos no código: quando mesclamos várias células, o 
Excel passa a tratar aquele bloco todo pelo nome da primeira célula (A1, no nosso caso). Porém, na hora 
de aplicar a cor de fundo via Python, pode ocorrer um bug visual onde apenas o primeiro espaço físico (a 
antiga A1) é pintado, deixando o resto do bloco mesclado em branco. Para sermos profissionais e 
evitarmos qualquer problema gráfico, usaremos uma técnica para garantir que todas as células que 
compõem o bloco mesclado recebam a mesma cor de fundo:</p>''')
st.code('''# --- Mesclagem e estilização do título principal --- #
ws.merge_cells('A1:D1')
celula_titulo = ws['A1']

# --- Aplicação da tipografia --- #
celula_titulo.font = Font(
    name='Calibri',
    size=16,
    bold=True,
    color='FFFFFF'
)

# --- Preenchimento do fundo da célula --- #
celula_titulo.fill = PatternFill(
    start_color='1F4E78',
    end_color='1F4E78',
    fill_type='solid'
)

# --- Alinhamento do texto --- #
celula_titulo.alignment = Alignment(
    horizontal='center',
    vertical='center'
)

# --- Pintar o fundo das demais células --- #
for coluna in range(2, 5):
    ws.cell(row=1, column=coluna).fill = PatternFill(
        start_color='1F4E78',
        end_color='1F4E78',
        fill_type='solid'
    )

wb.save('arquivo_02.xlsx')''', line_numbers=True)
st.html('''<p class="fonte_texto">Nesta etapa, o código foca inteiramente no impacto visual. Vamos 
analisar as ferramentas que utilizamos para dar esse salto de qualidade.</p>''')
st.html('''<p class="fonte_texto">A mágica começa com 
<span class="texto_python">ws.merge_cells(</span><span class="variaveis">'A1:D1'</span>
<span class="texto_python">)</span>. O nosso relatório tem 4 colunas de dados (ID, Produto, Quantidade e 
Receita), que ocupam as colunas A, B, C e D. Por isso, estamos dizendo ao Python para pegar a linha 1 
inteira, da coluna A até a D, e fundi-la em um único e espaçoso bloco.</p>''')
st.html('''<p class="fonte_texto">Logo abaixo, criamos a variável 
<span class="texto_python">celula_titulo = ws[</span><span class="variaveis">'A1'</span>
<span class="texto_python">]</span>. Como a célula A1 agora representa todo esse blocão mesclado, guardar 
isso em uma variável é uma excelente prática. Isso evita que tenhamos que digitar 
<span class="texto_python">ws[</span><span class="variaveis">'A1'</span>
<span class="texto_python">]</span> repetidamente nas próximas linhas, deixando o código limpo e 
elegante.</p>''')
st.html('''<p class="fonte_texto">No bloco de tipografia, chamamos a classe 
<span class="texto_python">Font</span> e passamos os atributos desejados para a nossa variável 
<span class="texto_python">celula_titulo</span>. Escolhemos a fonte 
<span class="variaveis">'Calibri'</span> (que já é nativa e segura para o Excel, evitando erros de 
compatibilidade), aumentamos o tamanho (<span class="texto_python">size=</span>
<span class="numeros">16</span>) para dar o destaque merecido a um título, e ativamos o negrito com 
<span class="texto_python">bold=</span><span class="palavras_reservadas">True</span>.</p>''')
st.html('''<p class="fonte_texto">Um ponto crucial é a cor (<span class="texto_python">color=</span>
<span class="variaveis">'FFFFFF'</span>). No 
<span class="texto_python">openpyxl</span>, trabalhamos com cores em formato hexadecimal, mas nunca 
utilizamos o símbolo de hashtag (<span class="texto_python">#</span>) antes do código, senão o programa 
retorna um erro. O código <span class="texto_python">FFFFFF</span> representa a cor branca, que fará um 
contraste perfeito com o fundo escuro que aplicaremos a seguir.</p>''')
st.html('''<p class="fonte_texto">Aqui é onde definimos a cor da nossa célula com a classe 
<span class="texto_python">PatternFill</span>. A estrutura exige que você defina uma 
<span class="texto_python">start_color</span> e uma <span class="texto_python">end_color</span>. Se você 
quisesse fazer um degradê, passaria cores diferentes, mas como queremos uma cor direta e chapada, 
repetimos o código <span class="texto_python">1F4E78</span> (um tom bem profissional de azul escuro) em 
ambos os parâmetros, e garantimos isso definindo <span class="texto_python">fill_type=</span>
<span class="variaveis">'solid'</span>.</p>''')
st.html('''<p class="fonte_texto">Não adianta ter uma célula gigante e bonita se o texto ficar espremido 
no canto esquerdo. Usando a classe <span class="texto_python">Alignment</span>, configuramos o 
alinhamento <span class="texto_python">horizontal=</span><span class="variaveis">'center'</span> e 
<span class="texto_python">vertical=</span><span class="variaveis">'center'</span>. Isso garante que o 
texto fique flutuando exatamente no meio do nosso painel azul, não importa o quanto a altura da linha ou 
a largura das colunas sejam ajustadas no futuro.</p>''')
st.html('''<p class="fonte_texto">Lembra do problema do bug visual que citei na teoria? É aqui que 
resolvemos isso como verdadeiros profissionais! Criamos um laço de repetição 
<span class="palavras_reservadas">for </span><span class="texto_python">coluna </span>
<span class="palavras_reservadas">in </span><span class="funcoes_python">range</span>
<span class="texto_python">(</span><span class="numeros">2</span><span class="texto_python">, </span>
<span class="numeros">5</span><span class="texto_python">)</span>:</p>''')
st.html('''<ul class="fonte_texto">
<li>O <span class="funcoes_python">range</span>
<span class="texto_python">(</span><span class="numeros">2</span><span class="texto_python">, </span>
<span class="numeros">5</span><span class="texto_python">)</span> gerará os números 2, 3 e 4. Ou seja, 
ele acessará as colunas B (2), C (3) e D (4). A coluna A (1) já foi pintada lá em cima.</li>
<li>Dentro do loop, nós navegamos obrigatoriamente pela linha 1 (
<span class="texto_python">row=</span><span class="numeros">1</span>) e aplicamos exatamente a mesma 
configuração de fundo azul que usamos no <span class="texto_python">PatternFill</span> anterior. O que 
estamos fazendo na prática? "Por baixo dos panos", estamos pintando as células que foram mescladas para 
garantir que a faixa azul cubra 100% da extensão da mescla, do início da coluna A até o final da coluna D, 
blindando nosso relatório contra qualquer falha gráfica do Excel.</li>
</ul>''')
st.html('''<p class="fonte_texto">Encerramos salvando tudo em 
<span class="texto_python">wb.save(</span><span class="variaveis">'arquivo_02.xlsx'</span>
<span class="texto_python">)</span>. Criar arquivos sequenciais é uma ótima estratégia de estudo e 
desenvolvimento. Assim, se você abrir o arquivo_01 e o arquivo_02 lado a lado, verá claramente o salto que 
demos da tabela crua para um título profissional e bem formatado!</p>''')
st.divider()

# --- Criação de cabeçalhos corporativos --- #
st.html('<h1 class="fonte_titulo_aula">Criação de cabeçalhos corporativos</h1>')
st.html('''<p class="fonte_texto">Nosso relatório já está ganhando uma cara totalmente nova com aquele 
título imponente que construímos! Mas, para que a nossa tabela converse visualmente e guie os olhos de 
quem a lê, os cabeçalhos das colunas (ID, Produto, Quantidade e Receita) também precisam de um tratamento 
especial. Afinal, eles são a porta de entrada para os nossos dados.</p>''')
st.html('''<p class="fonte_texto">A grande sacada teórica desta terceira etapa é a reutilização de código. 
Na parte anterior, aplicamos as formatações (fonte, cor, alinhamento) diretamente na célula do título. No 
entanto, como temos vários cabeçalhos, se fossemos aplicar a estilização célula por célula de forma manual, 
nosso código ficaria gigantesco e repetitivo. Para resolver isso com inteligência, criaremos verdadeiros 
"templates" (ou moldes) de estilos guardados em variáveis. Depois, com a ajuda de um laço de repetição, 
aplicaremos esse molde a todas as células do cabeçalho de uma só vez. É aqui que começamos a ver a 
introdução das bordas (<span class="texto_python">Border</span> e <span class="texto_python">Side</span>), 
um detalhe sutil, mas que faz toda a diferença na separação visual da tabela:</p>''')
st.code('''# --- Estilização dos cabeçalhos --- #
cabecalho_fundo = PatternFill(
    start_color='D9E1F2',
    end_color='D9E1F2',
    fill_type='solid'
)
cabecalho_fonte = Font(
    name='Calibri',
    size=11,
    bold=True,
    color='1F4E78'
)

# --- Configuração das bordas --- #
cabecalho_borda = Border(
    bottom=Side(style='medium', color='1F4E78'),
    top=Side(style='thin', color='D9D9D9'),
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9')
)

# --- Aplicação do estilo no cabeçalho --- #
for indice_coluna in range(1, 5):
    celula = ws.cell(row=2, column=indice_coluna)
    celula.fill = cabecalho_fundo
    celula.font = cabecalho_fonte
    celula.alignment = Alignment(horizontal='center', vertical='center')
    celula.border = cabecalho_borda

wb.save('arquivo_03.xlsx')''', line_numbers=True)
st.html('''<p class="fonte_texto">Essa parte do código é um excelente exemplo de como organizar a 
estilização de forma profissional e escalável. Vamos entender cada bloco.</p>''')
st.html('''<p class="fonte_texto">Em vez de aplicar diretamente, nós instanciamos as classes 
<span class="texto_python">PatternFill</span> e <span class="texto_python">Font</span> e as guardamos nas 
variáveis <span class="texto_python">cabecalho_fundo</span> e 
<span class="texto_python">cabecalho_fonte</span>.</p>''')
st.html('''<p class="fonte_texto">Escolhemos a cor hexadecimal 
<span class="texto_python">D9E1F2</span>, que é um tom de azul bem clarinho. Lembra que o nosso título 
tem um fundo azul escuro? Colocar um fundo mais claro nos cabeçalhos cria uma hierarquia visual super 
agradável e profissional.</p>''')
st.html('''<p class="fonte_texto">Mantemos a clássica 
<span class="variaveis">'Calibri'</span>, mas agora em tamanho 11 (menor que o título). Ativamos o negrito 
(<span class="texto_python">bold=</span><span class="palavras_reservadas">True</span>) e usamos a cor 
<span class="texto_python">1F4E78</span> (aquele azul escuro do título). Ou seja, invertemos a lógica do 
título: lá era fundo escuro com letra clara, aqui é fundo claro com letra escura.</p>''')
st.html('''<p class="fonte_texto">No Excel, uma célula tem quatro lados, e o 
<span class="texto_python">openpyxl</span> nos permite controlar cada um deles de forma independente. 
Para isso, usamos a classe <span class="texto_python">Border</span>, e dentro dela, definimos os lados (
<span class="texto_python">top</span>, <span class="texto_python">bottom</span>, 
<span class="texto_python">left</span>, <span class="texto_python">right</span>) usando a classe 
<span class="texto_python">Side</span>.</p>''')
st.html('''<p class="fonte_texto">Preste muita atenção na estratégia aqui: para o 
<span class="texto_python">bottom</span> (a linha de baixo do cabeçalho, que encosta nos dados), usamos o 
<span class="texto_python">style=</span><span class="variaveis">'medium'</span> (uma linha um pouco mais 
grossa) na cor azul escura. Isso cria uma barreira visual que diz "daqui para baixo, começam os 
dados".</p>''')
st.html('''<p class="fonte_texto">Para os demais lados, usamos o 
<span class="texto_python">style=</span><span class="variaveis">'thin'</span> (fino) em um tom de cinza 
clarinho (<span class="texto_python">D9D9D9</span>). Fica discreto, elegante e não polui a tela. Guardamos 
todo esse pacote na variável <span class="texto_python">cabecalho_borda</span>.</p>''')
st.html('''<p class="fonte_texto">Agora que temos nossos moldes de estilo prontos, precisamos aplicá-los 
nas colunas A, B, C e D da segunda linha (onde estão nossos cabeçalhos). Para isso, abrimos o loop 
<span class="palavras_reservadas">for </span><span class="texto_python">indice_coluna </span>
<span class="palavras_reservadas">in </span><span class="funcoes_python">range</span>
<span class="texto_python">(</span><span class="numeros">1</span>
<span class="texto_python">, </span><span class="numeros">5</span>
<span class="texto_python">)</span>:</p>''')
st.html('''<ul class="fonte_texto">
<li>Lembra que o <span class="funcoes_python">range</span>
<span class="texto_python">(</span><span class="numeros">1</span>
<span class="texto_python">, </span><span class="numeros">5</span>
<span class="texto_python">)</span> gera os números 1, 2, 3 e 4. Perfeito para acessar nossas quatro 
colunas.</li>
<li>Dentro do loop, criamos uma variável auxiliar temporária: 
<span class="texto_python">celula = ws.cell(row=</span><span class="numeros">2</span>
<span class="texto_python">, column=indice_coluna)</span>. Estamos travando a linha 2 (que é a linha dos 
cabeçalhos) e variando a coluna a cada rodada do loop.</li>
</ul>''')
st.html('''<p class="fonte_texto">Com a célula atual em mãos dentro do loop, o código fica lindo de ler. 
Nós simplesmente dizemos:</p>''')
st.html('''<ul class="fonte_texto">
<li><span class="texto_python">celula.fill = cabecalho_fundo</span>: aplique o molde de fundo.</li>
<li><span class="texto_python">celula.font = cabecalho_fonte</span>: aplique o molde de fonte.</li>
<li><span class="texto_python">celula.alignment = Alignment(...)</span>: centralize o texto, criando a 
classe diretamente aqui, pois é bem curta.</li>
<li><span class="texto_python">celula.border = cabecalho_borda</span>: aplique nosso pacote de bordas.</li>
</ul>''')
st.html('''<p class="fonte_texto">Por fim, rodamos o 
<span class="texto_python">wb.save(</span><span class="variaveis">'arquivo_03.xlsx'</span>
<span class="texto_python">)</span>. Ao abrir esse novo arquivo, você verá uma diferença gigantesca: a 
tabela já parece ter saído de um sistema corporativo profissional, com separações claras e cores 
harmônicas.</p>''')
st.divider()

# --- Alinhamento estrutural e formatação numérica de dados --- #
st.html('<h1 class="fonte_titulo_aula">Alinhamento estrutural e formatação numérica de dados</h1>')
st.html('''<p class="fonte_texto">Chegamos ao coração da nossa planilha: os dados! De nada adianta termos 
um título belíssimo e cabeçalhos bem definidos se os números e textos da nossa tabela estiverem 
bagunçados. Quando lidamos com relatórios, a forma como a informação é apresentada faz toda a diferença na 
facilidade de leitura. Números jogados de qualquer jeito, sem separador de milhar ou formatação de moeda, 
confundem o usuário e tiram todo o aspecto profissional do seu trabalho.</p>''')
st.html('''<p class="fonte_texto">A teoria por trás dessa etapa gira em torno de duas regras de ouro do 
design de planilhas. A primeira é o alinhamento lógico: textos (como o nome do produto) devem ser 
alinhados à esquerda para facilitar a leitura ocidental; números e moedas devem ser alinhados à direita, 
para que as unidades, dezenas e centenas fiquem perfeitamente empilhadas; e identificadores (como o ID) 
ficam ótimos centralizados. A segunda regra é a máscara de formatação (
<span class="texto_python">number_format</span>). A grande mágica aqui é que não vamos transformar nossos 
números em texto (strings) apenas para colocar um "R$" na frente. Se fizéssemos isso, o Excel não 
conseguiria somar esses valores depois. Nós aplicaremos uma máscara visual: o Python dirá ao Excel 
para exibir o R$ e os separadores, mas manter o valor real matemático intacto por baixo dos panos.</p>''')
st.code('''# --- Formatação visual das colunas de dados --- #
lado_fino = Side(style='thin', color='D9D9D9')
dados_borda = Border(
    top=lado_fino,
    bottom=lado_fino,
    left=lado_fino,
    right=lado_fino
)

for linha in range(3, 7):
    # --- Coluna 1: ID do produto (alinhamento centralizado e bordas padrão) --- #
    celula_id = ws.cell(row=linha, column=1)
    celula_id.alignment = Alignment(horizontal='center')
    celula_id.border = dados_borda

    # --- Coluna 2: Produto (alinhamento à esquerda para melhor legibilidade textual) --- #
    celula_produto = ws.cell(row=linha, column=2)
    celula_produto.alignment = Alignment(horizontal='left')
    celula_produto.border = dados_borda

    # --- Coluna 3: Quantidade (alinhamento à direita e separação de milhar) --- #
    celula_qtd = ws.cell(row=linha, column=3)
    celula_qtd.alignment = Alignment(horizontal='right')
    celula_qtd.number_format = '#,##0'
    celula_qtd.border = dados_borda

    # --- Coluna 4: Receita (alinhamento à direita e formatação monetária brasileira) --- #
    celula_receita = ws.cell(row=linha, column=4)
    celula_receita.alignment = Alignment(horizontal='right')
    celula_receita.number_format = 'R$ #,##.00'
    celula_receita.border = dados_borda

wb.save('arquivo_04.xlsx')''', line_numbers=True)
st.html('''<p class="fonte_texto">Essa é a parte em que a nossa tabela passa de um simples rascunho para 
um documento pronto para ser entregue a uma diretoria. Vamos dissecar a lógica utilizada.</p>''')
st.html('''<p class="fonte_texto">Logo no início, nós criamos um padrão de borda bem discreto para 
envolver todas as células de dados. Instanciamos um <span class="texto_python">Side(style=</span>
<span class="variaveis">'thin'</span><span class="texto_python">, color=</span>
<span class="variaveis">'D9D9D9'</span><span class="texto_python">)</span> (linha fina e cinza clara) e 
passamos esse mesmo lado para o topo, base, esquerda e direita dentro da classe Border. Salvamos tudo na 
variável <span class="texto_python">dados_borda</span>. Ter esse molde pronto fora do loop economiza 
processamento e dezenas de linhas de código.</p>''')
st.html('''<p class="fonte_texto">Para formatar os dados, abrimos o loop 
<span class="palavras_reservadas">for </span><span class="texto_python">linha </span>
<span class="palavras_reservadas">in </span><span class="funcoes_python">range</span>
<span class="texto_python">(</span><span class="numeros">3</span>
<span class="texto_python">, </span><span class="numeros">7</span><span class="texto_python">)</span>. 
Por que de 3 a 7? Porque a linha 1 é o nosso título e a linha 2 é o nosso cabeçalho. Nossos dados 
começam na linha 3 e terminam na linha 6 (lembre-se que a função range no Python para um número antes do 
final, então <span class="funcoes_python">range</span>
<span class="texto_python">(</span><span class="numeros">3</span>
<span class="texto_python">, </span><span class="numeros">7</span><span class="texto_python">)</span> 
processará as linhas 3, 4, 5 e 6). Dentro deste loop, nós vamos "pescar" cada célula individualmente e 
dar o tratamento adequado para a sua respectiva coluna.</p>''')
st.html('''<p class="fonte_texto">Na Coluna 1 (ID), capturamos a célula com 
<span class="texto_python">ws.cell(row=linha, column=</span><span class="numeros">1</span>
<span class="texto_python">)</span>. Como IDs numéricos não são usados para contas matemáticas e servem 
apenas como códigos, aplicamos o <span class="texto_python">Alignment(horizontal=</span>
<span class="variaveis">'center'</span><span class="texto_python">)</span> para centralizá-los sob o 
cabeçalho e adicionamos a nossa borda padrão.''')
st.html('''<p class="fonte_texto">Na Coluna 2 (Produto), temos os nomes dos produtos. Para blocos de 
texto, a leitura flui muito melhor quando estão ancorados à esquerda. Por isso, usamos 
<span class="texto_python">Alignment(horizontal=</span><span class="variaveis">'left'</span>
<span class="texto_python">)</span> e aplicamos a mesma borda.''')
st.html('''<p class="fonte_texto">Aqui a brincadeira começa a ficar interessante. Capturamos a Coluna 3 
(Quantidade) e a alinhamos à direita. O "pulo do gato" está na propriedade 
<span class="texto_python">number_format = </span><span class="variaveis">'#,##0'</span>. Esse código 
interno diz ao Excel: <i>"Se o número passar de mil, coloque um separador de milhar"</i>. Note que, no 
padrão de formatação internacional do Excel via código, a vírgula (<span class="texto_python">,</span>) 
representa a separação de milhar. Quando você abrir isso no seu Excel em português, ele automaticamente 
traduzirá e exibirá com pontos, deixando visualmente impecável.''')
st.html('''<p class="fonte_texto">A nossa Coluna 4 (Receita) é onde está o dinheiro, então ela merece 
brilhar. Alinhamos à direita e aplicamos a propriedade <span class="texto_python">number_format = </span>
<span class="variaveis">'R$ #,##.00'</span>. Preste muita atenção na sintaxe desta string:''')
st.html('''<ul class="fonte_texto">
<li>O <span class="texto_python">R$</span> é inserido literalmente, com um espaço, para garantir que o 
símbolo da moeda apareça de forma limpa.</li>
<li>O <span class="texto_python">#,##</span> aplica a quebra de milhar.</li>
<li>O <span class="texto_python">.00</span> força o Excel a exibir sempre duas casas decimais, preenchendo 
com zeros se o número for redondo. No Excel brasileiro, esse ponto será automaticamente convertido para a 
vírgula dos centavos. E, claro, finalizamos aplicando a 
<span class="texto_python">dados_borda</span> para fechar a caixinha perfeitamente.</li>
</ul>''')
st.html('''<p class="fonte_texto">Executamos o 
<span class="texto_python">wb.save(</span><span class="variaveis">'arquivo_04.xlsx'</span>
<span class="texto_python">)</span> para consolidar essas alterações em um novo arquivo. Ao comparar o 
arquivo_03 com este arquivo_04, o salto de maturidade dos dados será nítido: tudo alinhado, separado e 
formatado como dinheiro de verdade.''')
st.divider()

# --- Dimensionamento dinâmico das colunas --- #
st.html('<h1 class="fonte_titulo_aula">Dimensionamento dinâmico das colunas</h1>')
st.html('''<p class="fonte_texto">Nossa tabela está ficando com um visual cada vez mais refinado, mas 
agora nos deparamos com aquele problema clássico e irritante do Excel: os famigerados "hashtags" (
<span class="texto_python">###</span>). Isso acontece porque, até o momento, ajustamos o conteúdo 
interno das células, mas a largura da coluna em si continua com o tamanho padrão. Quando o Excel percebe 
que um número não cabe no espaço físico da célula, ele o esconde atrás de ### para evitar que você leia 
uma informação cortada (e potencialmente errada, como ler "100" em vez de "1000"). Textos maiores, como 
"Monitor Ultrawide", também ficam espremidos ou sobrepostos.''')
st.html('''<p class="fonte_texto">A solução "braçal" seria clicar e arrastar a largura de cada coluna lá 
no próprio Excel. Mas estamos programando em Python justamente para não ter que fazer trabalho braçal! A 
teoria brilhante desta etapa é a construção de um algoritmo de dimensionamento dinâmico. Vamos criar um 
laço de repetição que varrerá a nossa planilha de cima a baixo, coluna por coluna. Para cada coluna, 
ele olhará todas as células e contar quantos caracteres existem ali. O Python guardará o tamanho da 
"maior palavra" que encontrar e, em seguida, usará esse valor máximo para ajustar a largura de toda a 
coluna de forma automática. Assim, a tabela vai "respirar" e se expandir perfeitamente para acomodar 
qualquer dado que colocarmos nela:''')
st.code('''# --- Redimensionar dinamicamente a largura das colunas --- #
for coluna in ws.columns:
    tamanho_max = 0
    letra_coluna = get_column_letter(coluna[0].column)
    for celula in coluna:
        # --- Avaliar a string da célula para mensuração física --- #
        valor = str(celula.value or '')
        if len(valor) > tamanho_max:
            tamanho_max = len(valor)

    # --- Atribuição da largura ideal --- #
    ws.column_dimensions[letra_coluna].width = max(tamanho_max + 5, 10)

wb.save('arquivo_05.xlsx')''', line_numbers=True)
st.html('''<p class="fonte_texto">Esse é, sem dúvida, o bloco mais lógico e "programático" que fizemos 
até agora. Vamos entender como esse algoritmo varre a planilha e resolve o nosso problema de espaçamento.''')
st.html('''<p class="fonte_texto">Abrimos o loop principal com 
<span class="palavras_reservadas">for </span><span class="texto_python">coluna </span>
<span class="palavras_reservadas">in </span><span class="texto_python">ws.columns</span>. O atributo 
<span class="texto_python">ws.columns</span> pega a nossa aba ativa e gera um pacote contendo todas as 
colunas que possuem dados (no nosso caso, as colunas A, B, C e D). A cada rodada do loop, ele nos entrega 
uma coluna inteira para trabalharmos.''')
st.html('''<p class="fonte_texto">Dentro desse loop, zeramos a variável 
<span class="texto_python">tamanho_max = </span><span class="numeros">0</span>. Esse é o nosso contador, 
e ele precisa começar do zero toda vez que mudarmos para uma nova coluna.''')
st.html('''<p class="fonte_texto">Em seguida, usamos a fantástica função 
<span class="texto_python">get_column_letter()</span>, que importamos lá no início do código. Passamos 
para ela <span class="texto_python">coluna[</span><span class="numeros">0</span>
<span class="texto_python">].column</span> (que pega o número da primeira célula dessa coluna) e ela nos 
devolve a letra correspondente! Se estivermos na coluna 1, ela nos devolve 'A'; se na 2, devolve 'B', e 
assim por diante. Guardamos isso na variável <span class="texto_python">letra_coluna</span>, pois 
precisaremos dessa letra para redimensionar o Excel mais à frente.''')
st.html('''<p class="fonte_texto">Agora abrimos um segundo loop, <i>dentro</i> do primeiro: 
<span class="palavras_reservadas">for </span><span class="texto_python">celula </span>
<span class="palavras_reservadas">in </span><span class="texto_python">coluna</span>. Aqui nós vamos 
descer linha por linha dentro da coluna atual (ex: A1, A2, A3, A4...).''')
st.html('''<p class="fonte_texto">A linha 
<span class="texto_python">valor = </span><span class="funcoes_python">str</span>
<span class="texto_python">(celula.value </span><span class="palavras_reservadas">or </span>
<span class="variaveis">''</span><span class="texto_python">)</span> é um show de eficiência em 
Python:''')
st.html('''<ul class="fonte_texto">
<li><span class="texto_python">celula.value</span> pega o conteúdo da célula.</li>
<li>O <span class="palavras_reservadas">or </span>
<span class="variaveis">''</span> (ou aspas vazias) é uma trava de segurança: se a célula estiver em 
branco (<span class="palavras_reservadas">None</span>), ele transforma em uma string vazia em vez de 
dar erro.</li>
<li>Por fim, o <span class="funcoes_python">str</span><span class="texto_python">()</span> força a 
conversão de tudo (inclusive números inteiros e decimais) para texto. Precisamos fazer isso para conseguir 
contar quantos caracteres existem ali.</li>
</ul>''')
st.html('''<p class="fonte_texto">O bloco <span class="palavras_reservadas">if </span>
<span class="funcoes_python">len</span><span class="texto_python">(valor) > tamanho_max</span> faz a 
checagem lógica. Usamos <span class="funcoes_python">len</span><span class="texto_python">()</span> para 
descobrir o comprimento (length) da palavra atual. Se a palavra na célula A1 tem 5 letras, o nosso 
<span class="texto_python">tamanho_max</span> (que era 0) passa a valer 5. Se na célula A2 tiver uma 
palavra com 12 letras, o <span class="texto_python">tamanho_max</span> é atualizado para 12. No final 
dessa inspeção, <span class="texto_python">tamanho_max</span> terá o tamanho exato do maior dado presente 
naquela coluna.''')
st.html('''<p class="fonte_texto">Fora do loop das células, mas ainda dentro do loop das colunas, nós 
finalmente mudamos a estrutura do Excel com o comando <span class="texto_python">ws.column_dimensions
[letra_coluna].width = </span><span class="palavras_python">max</span>
<span class="texto_python">(tamanho_max + </span><span class="numeros">5</span>
<span class="texto_python">, </span><span class="numeros">10</span>
<span class="texto_python">)</span>.''')
st.html('''<p class="fonte_texto"><span class="texto_python">ws.column_dimensions[</span>
<span class="variaveis">'A'</span><span class="texto_python">].width</span> é como pegamos a dimensão da 
coluna A para alterá-la.''')
st.html('''<p class="fonte_texto">Usamos a função matemática nativa 
<span class="funcoes_python">max</span><span class="texto_python">()</span> para definir a largura. Por 
que fazemos isso? A nossa fórmula é <span class="texto_python">tamanho_max + </span>
<span class="numeros">5</span>. O "+ 5" funciona como um respiro, uma margem de segurança para o texto não 
ficar colado na borda.''')
st.html('''<p class="fonte_texto">O <span class="funcoes_python">max</span>
<span class="texto_python">(valor1, valor2)</span> escolhe o maior número entre as opções. Ou seja, se o 
maior texto da coluna tiver tamanho 2, a fórmula ficaria <span class="funcoes_python">max</span>
<span class="texto_python">(</span><span class="numeros">2 </span><span class="texto_python">+ </span>
<span class="numeros">5</span><span class="texto_python">, </span>
<span class="numeros">10</span><span class="texto_python">)</span>. Como 10 é maior que 7, a largura da 
coluna será fixada em 10. Isso garante que as colunas nunca fiquem minúsculas demais, mantendo um tamanho 
mínimo padrão de 10 unidades de largura.''')
st.html('''<p class="fonte_texto">Executamos o <span class="texto_python">wb.save(</span>
<span class="variaveis">'arquivo_05.xlsx'</span><span class="texto_python">)</span>. Ao abri-lo, você 
verá que aquelas barras de hashtags sumiram e textos longos como "Mouse Ergonômico" estão perfeitamente 
legíveis, com a coluna B esticada na medida exata.''')
st.divider()

# --- Otimização visual da área de trabalho --- #
st.html('<h1 class="fonte_titulo_aula">Otimização visual da área de trabalho</h1>')
st.html('''<p class="fonte_texto">Chegamos à reta final do nosso projeto! Nossa tabela já está com um 
design incrível, os dados estão perfeitamente formatados como moedas e separadores de milhar, e as colunas 
se ajustaram magicamente ao tamanho do texto. Qualquer pessoa que visse a tabela agora já ficaria 
impressionada. No entanto, para entregarmos um material verdadeiramente impecável, precisamos aplicar a 
"cereja do bolo". Nesta sexta e última etapa, o nosso foco não é mais alterar os dados em si, mas sim 
manipular a <b>visualização da área de trabalho</b> do Excel.''')
st.html('''<p class="fonte_texto">A parte teórica aqui envolve a Experiência do Usuário (UX) dentro da 
planilha. Quando abrimos o Excel, por padrão, ele nos mostra aquela infinidade de linhas e colunas cinzas 
(as linhas de grade) que preenchem a tela inteira. Para um relatório, isso gera muita poluição visual. O 
objetivo é transformar a planilha em uma tela em branco, como um "dashboard" de um sistema, onde apenas a 
nossa tabela desenhada chame a atenção. Além disso, precisamos garantir a navegabilidade: se a nossa 
tabela tivesse mil linhas, ao rolar o mouse para baixo, o usuário perderia o cabeçalho de vista e não 
saberia mais o que cada coluna significa. Congelar os painéis resolve esse problema, mantendo o título e 
os cabeçalhos fixos no topo, independentemente do quanto a pessoa role a tela para baixo:''')
st.code('''# --- Ocultação das linhas de grade --- #
ws.sheet_view.showGridLines = False

# --- Projetar o foco visual para a célula de origem (A1) para impedir corrupção --- #
ws.sheet_view.topleftCell = 'A1'

# --- Congelar as primeiras duas linhas --- #
ws.freeze_panes = 'A3'

# --- Salvar o arquivo final --- #
wb.save('faturamento.xlsx')''', line_numbers=True)
st.html('''<p class="fonte_texto">Esses últimos comandos são curtos, mas geram um impacto visual e de 
usabilidade gigantesco no resultado final. Vamos destrinchar cada um deles.''')
st.html('''<p class="fonte_texto">A linha <span class="texto_python">ws.sheet_view.showGridLines = </span>
<span class="palavras_reservadas">False</span> é o nosso limpador de tela. O atributo 
<span class="texto_python">sheet_view</span> controla como a aba é exibida na tela do computador. Ao 
definir showGridLines como <span class="palavras_reservadas">False</span>, nós desligamos aquela grade 
cinza padrão do Excel. O resultado? Todo o espaço em volta da nossa tabela fica inteiramente branco, 
dando um destaque absoluto e muito profissional ao quadro de faturamento que acabamos de pintar e 
estilizar. Tudo o que não for a nossa tabela deixa de competir pela atenção do usuário.''')
st.html('''<p class="fonte_texto">No comando 
<span class="texto_python">ws.sheet_view.topleftCell = </span><span class="variaveis">'A1'</span>, 
estamos forçando o Excel a posicionar a "câmera" no canto superior esquerdo assim que o arquivo for 
aberto. Por que isso é importante? Muitas vezes, durante a automação ou quando salvamos um arquivo, o 
foco da planilha pode acabar ficando perdido lá na linha 500 ou em alguma coluna distante, o que causa um 
desconforto (ou até uma leve "corrupção visual") para quem abre o relatório, achando que ele está em 
branco ou quebrado. Com essa linha, nós garantimos que a primeira coisa que o seu chefe ou cliente vai 
ver ao dar dois cliques no arquivo é o nosso belíssimo título na célula A1.''')
st.html('''<p class="fonte_texto">Aqui temos um dos recursos mais úteis do Excel feito via Python: 
<span class="texto_python">ws.freeze_panes = </span><span class="variaveis">'A3'</span>. A lógica de 
congelamento do <span class="texto_python">openpyxl</span> (e do próprio Excel) funciona da seguinte 
maneira: ele congela tudo o que está acima e à esquerda da célula que você especificou.''')
st.html('''<p class="fonte_texto">Como escolhemos a célula A3 (onde os nossos dados começam), o Excel 
congela tudo o que está acima dela, ou seja, as linhas 1 e 2.''')
st.html('''<p class="fonte_texto">Como a célula está na coluna A, não há nada à esquerda dela para 
congelar. O resultado prático disso é que o nosso Título (linha 1) e o nosso Cabeçalho (linha 2) ficam 
travados na tela. O usuário pode descer mil linhas de dados de faturamento que o cabeçalho sempre estará 
ali em cima, servindo como guia.''')
st.html('''<p class="fonte_texto">Finalmente, substituímos os nossos arquivos de rascunho temporários 
pelo comando <span class="texto_python">wb.save(</span>
<span class="variaveis">'faturamento.xlsx'</span><span class="texto_python">)</span>. Agora estamos 
salvando o nosso projeto com o nome oficial e definitivo.''')
st.divider()

# --- Resumo --- #
st.html('<h1 class="fonte_titulo_aula">Resumo</h1>')
st.html('''<p class="fonte_texto">Nesta terceira aula, demos um salto gigantesco ao transformar dados 
brutos em um relatório de faturamento com padrão corporativo utilizando a biblioteca 
<span class="texto_python">openpyxl</span>. O passo inicial focou na estruturação da nossa base: criamos 
um <i>Workbook</i> do zero, renomeamos a aba de trabalho e descarregamos nossa matriz de informações com 
loops inteligentes. Essa fundação foi essencial para preparar o terreno para a verdadeira mágica da 
automação, garantindo que cada linha e coluna estivesse em seu devido lugar antes de receber qualquer 
tratamento estético.''')
st.html('''<p class="fonte_texto">Com a estrutura montada, mergulhamos fundo na estilização profissional. 
Aprendemos a mesclar células para construir um painel de título imponente e utilizamos ferramentas 
poderosas como <span class="texto_python">Font</span>, <span class="texto_python">PatternFill</span>, 
<span class="texto_python">Alignment</span> e Border para aplicar uma identidade visual coesa e 
elegante. Além da paleta de cores e tipografia, aplicamos formatações lógicas aos dados: alinhamos textos 
à esquerda para facilitar a leitura e inserimos máscaras financeiras completas (R$ e separadores de 
milhar) nos números. O grande trunfo dessa formatação é que ela apenas altera a visualização, mantendo os 
valores intactos por baixo dos panos para que o Excel ainda consiga realizar cálculos matemáticos com 
eles.''')
st.html('''<p class="fonte_texto">Para coroar o projeto, implementamos soluções avançadas de layout e 
experiência do usuário (UX). Construímos um algoritmo inteligente de dimensionamento dinâmico que varre a 
planilha, calcula o tamanho do maior texto de cada coluna e ajusta a largura perfeitamente, banindo as 
temidas hashtags (<span class="texto_python">###</span>) da tela. Por fim, limpamos a poluição visual 
ocultando as linhas de grade padrão (<span class="texto_python">showGridLines = </span>
<span class="palavras_reservadas">False</span>) e travamos o cabeçalho no topo (
<span class="texto_python">freeze_panes</span>), garantindo que a navegação continue intuitiva mesmo em 
tabelas gigantescas. O resultado final é um script reaproveitável que gera planilhas de nível executivo 
com apenas um clique.''')
st.divider()

# --- Conclusão --- #
st.html('<h1 class="fonte_titulo_aula">Conclusão</h1>')
st.html('''<p class="fonte_texto">Chegamos ao fim da nossa terceira etapa com um resultado que fala por 
si só! O que começou como uma simples inserção de dados em uma planilha crua se transformou em um relatório 
corporativo elegante, funcional e pronto para ser apresentado a qualquer diretoria. Vimos que o poder do 
Python, aliado à biblioteca <span class="texto_python">openpyxl</span>, vai muito além de apenas "jogar" 
informações em células. Nós assumimos o controle total do design: aplicamos cores sólidas, fontes 
personalizadas, bordas estratégicas e formatações financeiras que tornam a leitura agradável, tudo isso 
sem quebrar a inteligência matemática do Excel para cálculos futuros.''')
st.html('''<p class="fonte_texto">A grande lição que fica desta aula é o verdadeiro impacto da automação 
no seu dia a dia. Sim, estruturar todo esse código exigiu dedicação e atenção aos detalhes, mas pense no 
tempo que você economizou para o seu "eu do futuro"! Agora você tem em mãos um script robusto que não 
apenas preenche dados, mas redimensiona larguras de colunas automaticamente, oculta linhas de grade e 
congela painéis sem que você precise dar um único clique no mouse. Seja para uma tabela com cinco 
produtos ou para um banco de dados com cem mil linhas, essa exata mesma estrutura fará o trabalho visual 
pesado em frações de segundo, livrando você para sempre da formatação manual.''')
st.html('''<p class="fonte_texto">Espero que esse mergulho na estilização avançada tenha aberto a sua 
mente para as infinitas possibilidades que o Python oferece. Salve esse código, use-o como o seu gabarito 
oficial e não tenha medo de alterá-lo para testar as paletas de cores da sua própria empresa. Continue 
praticando e aplicando essa lógica nos seus projetos pessoais e profissionais. Nos vemos no próximo 
conteúdo para darmos mais um passo na nossa jornada de automação.''')
st.subheader('No mais é isso, nos vemos na próxima aula! Até lá, fiquem com Deus e fui!')